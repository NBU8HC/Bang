import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from datetime import datetime
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")

def extract_all_parameters_from_file(file_path):
    """Extract all parameters from a DCM file efficiently."""
    parameters = []
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    try:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as file:
                content = file.read()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return parameters
    
    lines = content.split('\n')
    i = 0
    
    while i < len(lines):
        stripped = lines[i].strip()
        
        # Check if this line starts a parameter definition
        if stripped.startswith(keywords):
            parts = stripped.split()
            if len(parts) > 1:
                param_name = parts[1]
                param_type = parts[0]
                
                # Collect the entire block until END
                block_lines = [lines[i]]
                i += 1
                
                while i < len(lines):
                    block_lines.append(lines[i])
                    if lines[i].strip() == 'END':
                        break
                    i += 1
                
                block = '\n'.join(block_lines)
                
                # Extract value from block
                param_value = ""
                for line in block_lines:
                    line_stripped = line.strip()
                    if line_stripped.startswith('WERT') or line_stripped.startswith('TEXT'):
                        value_parts = line_stripped.split(maxsplit=1)
                        if len(value_parts) > 1:
                            param_value = value_parts[1]
                        break
                
                # Add to parameters list
                parameters.append({
                    'name': param_name,
                    'type': param_type,
                    'value': param_value,
                    'block': block
                })
        
        i += 1
    
    return parameters

def process_multiple_directories_parallel(directories, config_parameters, file_extension=".dcm", max_workers=8, track_updates=False):
    """Song song: tìm + cập nhật các file trong nhiều thư mục."""
    def work(d):
        matching = find_files_with_parameters(d, config_parameters, file_extension)
        if track_updates:
            result = update_files_in_directory(d, config_parameters, file_extension, track_updates=True)
            updated, update_tracking = result
            return d, matching, updated, update_tracking
        else:
            updated = update_files_in_directory(d, config_parameters, file_extension, track_updates=False)
            return d, matching, updated, []

    all_matching = []
    all_updated = []
    all_tracking = []
    per_dir = []
    if not directories:
        if track_updates:
            return all_matching, all_updated, per_dir, all_tracking
        return all_matching, all_updated, per_dir
    
    workers = min(max_workers, len(directories))
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futures = {ex.submit(work, d): d for d in directories}
        for fut in as_completed(futures):
            d = futures[fut]
            try:
                dir_path, matching, updated, tracking = fut.result()
                per_dir.append((dir_path, matching, updated))
                all_matching.extend(matching)
                all_updated.extend(updated)
                all_tracking.extend(tracking)
            except Exception as e:
                print(f"Error processing directory {d}: {e}")
                per_dir.append((d, [], []))
    
    if track_updates:
        return all_matching, all_updated, per_dir, all_tracking
    return all_matching, all_updated, per_dir

def read_config_file(config_path):
    """Read the config file and extract parameters and their values."""
    parameters = {}
    try:
        try:
            with open(config_path, 'r', encoding='utf-8') as file:
                lines = file.readlines()
        except UnicodeDecodeError:
            with open(config_path, 'r', encoding='latin-1') as file:
                lines = file.readlines()
    except FileNotFoundError:
        raise FileNotFoundError(f"Config file not found: {config_path}")
    except Exception as e:
        raise Exception(f"Error reading config file: {e}")
    
    current_param = None
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    for line_num, line in enumerate(lines, 1):
        line = line.strip()
        if line.startswith(keywords):
            parts = line.split()
            if len(parts) > 1:
                current_param = parts[1]
            else:
                print(f"Line {line_num}: Invalid parameter format: {line}")
        elif line.startswith('WERT') and current_param:
            parts = line.split()
            if len(parts) > 1:
                parameters[current_param] = parts[1]
            else:
                print(f"Line {line_num}: Invalid WERT format: {line}")
            current_param = None
        elif line.startswith('TEXT') and current_param:
            parts = line.split()
            if len(parts) > 1:
                parameters[current_param] = parts[1]
            else:
                print(f"Line {line_num}: Invalid TEXT format: {line}")
            current_param = None
    
    return parameters

def read_file(file_path):
    """Utility function to read file contents safely."""
    try:
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read()
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as file:
                return file.read()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return ""

def find_files_with_parameters(directory, parameters, file_extension=".dcm"):
    """Find files in the directory (recursively) containing any of the specified parameters."""
    matching_files = []
    parameter_occurrences = {}
    
    # Recursively find all files with the specified extension
    all_files = []
    for root, dirs, files in os.walk(directory):
        for filename in files:
            if filename.endswith(file_extension):
                file_path = os.path.join(root, filename)
                all_files.append(file_path)
    
    for file_path in all_files:
        filename = os.path.basename(file_path)
        try:
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    lines = file.readlines()
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='latin-1') as file:
                    lines = file.readlines()
        except Exception as e:
            print(f"Error reading file {filename}: {e}")
            continue
            
        found_params = []
        current_param = None
        keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
        
        for line_num, line in enumerate(lines, 1):
            stripped_line = line.strip()
            if stripped_line.startswith(keywords):
                parts = stripped_line.split()
                if len(parts) > 1:
                    current_param = parts[1]
                    if current_param in parameters:
                        found_params.append(current_param)
        
        if found_params:
            matching_files.append(file_path)
            parameter_occurrences[file_path] = list(set(found_params))

    return matching_files

def update_files_in_directory(directory, config_parameters, file_extension=".dcm", track_updates=False):
    """Update parameters in files within the specified directory (recursively)."""
    updated_files = []
    update_tracking = []  # Track individual updates
    
    # Recursively find all files with the specified extension
    for root, dirs, files in os.walk(directory):
        for filename in files:
            if filename.endswith(file_extension):
                file_path = os.path.join(root, filename)
            
                try:
                    try:
                        with open(file_path, 'r', encoding='utf-8') as file:
                            lines = file.readlines()
                        encoding_used = 'utf-8'
                    except UnicodeDecodeError:
                        with open(file_path, 'r', encoding='latin-1') as file:
                            lines = file.readlines()
                        encoding_used = 'latin-1'
                except Exception as e:
                    print(f"Error reading file {filename}: {e}")
                    continue

                updated_lines = []
                current_param = None
                current_block_lines = []
                in_block = False
                keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
                file_was_modified = False
                
                for i, line in enumerate(lines):
                    stripped_line = line.strip()
                    
                    if stripped_line.startswith(keywords):
                        parts = stripped_line.split()
                        if len(parts) > 1:
                            current_param = parts[1]
                            in_block = True
                            current_block_lines = [line]
                            updated_lines.append(line)
                            continue
                    
                    if in_block:
                        current_block_lines.append(line)
                        
                        if stripped_line == 'END':
                            in_block = False
                            if current_param not in config_parameters:
                                current_param = None
                                current_block_lines = []
                    
                    if stripped_line.startswith('WERT') and current_param and current_param in config_parameters:
                        # Capture old block before modification
                        old_block = ''.join(current_block_lines)
                        
                        # Extract old value
                        old_value = stripped_line.split(maxsplit=1)[1] if len(stripped_line.split()) > 1 else ""
                        new_value = config_parameters[current_param]
                        indentation = line[:len(line) - len(line.lstrip())]
                        line = f"{indentation}WERT {new_value}\n"
                        file_was_modified = True
                        
                        # Build new block (replace WERT line in current block)
                        new_block_lines = []
                        for bl in current_block_lines[:-1]:  # Exclude current WERT line
                            new_block_lines.append(bl)
                        new_block_lines.append(line)
                        
                        # Track update with full blocks
                        if track_updates:
                            update_tracking.append({
                                'file': file_path,
                                'folder': directory,
                                'parameter': current_param,
                                'old_value': old_block.rstrip(),
                                'new_value': '',  # Will be built when we hit END
                                'status': 'Updated',
                                '_building_new': new_block_lines
                            })
                    
                    elif stripped_line.startswith('TEXT') and current_param and current_param in config_parameters:
                        # Capture old block before modification
                        old_block = ''.join(current_block_lines)
                        
                        # Extract old value
                        old_value = stripped_line.split(maxsplit=1)[1] if len(stripped_line.split()) > 1 else ""
                        new_value = config_parameters[current_param]
                        indentation = line[:len(line) - len(line.lstrip())]
                        line = f"{indentation}TEXT {new_value}\n"
                        file_was_modified = True
                        
                        # Build new block (replace TEXT line in current block)
                        new_block_lines = []
                        for bl in current_block_lines[:-1]:  # Exclude current TEXT line
                            new_block_lines.append(bl)
                        new_block_lines.append(line)
                        
                        # Track update with full blocks
                        if track_updates:
                            update_tracking.append({
                                'file': file_path,
                                'folder': directory,
                                'parameter': current_param,
                                'old_value': old_block.rstrip(),
                                'new_value': '',  # Will be built when we hit END
                                'status': 'Updated',
                                '_building_new': new_block_lines
                            })
                    
                    updated_lines.append(line)
                    
                    # When we hit END, complete the new_value block in tracking
                    if stripped_line == 'END' and track_updates and update_tracking:
                        for track_entry in reversed(update_tracking):
                            if '_building_new' in track_entry and not track_entry['new_value']:
                                track_entry['_building_new'].append(line)
                                track_entry['new_value'] = ''.join(track_entry['_building_new']).rstrip()
                                del track_entry['_building_new']
                                break
                        current_param = None
                        current_block_lines = []
                
                if file_was_modified:
                    try:
                        with open(file_path, 'w', encoding= encoding_used) as file:
                            file.writelines(updated_lines)
                        updated_files.append(file_path)
                    except Exception as e:
                        print(f"Error writing to file {filename}: {e}")
    
    if track_updates:
        return updated_files, update_tracking
    return updated_files

def update_specific_files_optimized(file_cache_dict, config_parameters, track_updates=False):
    """Update specific files using cached content for maximum speed."""
    updated_files = []
    update_tracking = []
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    for file_path, cache_data in file_cache_dict.items():
        content = cache_data['content']
        encoding_used = cache_data['encoding']
        lines = content.split('\\n')
        
        # Convert back to line format with newlines
        lines = [line + '\\n' for line in lines]
        if lines and not lines[-1].endswith('\\n'):
            lines[-1] = lines[-1].rstrip('\\n')

        updated_lines = []
        current_param = None
        current_block_lines = []
        in_block = False
        file_was_modified = False
        
        for i, line in enumerate(lines):
            stripped_line = line.strip()
            
            if stripped_line.startswith(keywords):
                parts = stripped_line.split()
                if len(parts) > 1:
                    current_param = parts[1]
                    in_block = True
                    current_block_lines = [line]
                    updated_lines.append(line)
                    continue
            
            if in_block:
                current_block_lines.append(line)
                
                if stripped_line == 'END':
                    in_block = False
                    if current_param not in config_parameters:
                        current_param = None
                        current_block_lines = []
            
            if stripped_line.startswith('WERT') and current_param and current_param in config_parameters:
                old_block = ''.join(current_block_lines)
                new_value = config_parameters[current_param]
                indentation = line[:len(line) - len(line.lstrip())]
                line = f"{indentation}WERT {new_value}\\n"
                file_was_modified = True
                
                new_block_lines = []
                for bl in current_block_lines[:-1]:
                    new_block_lines.append(bl)
                new_block_lines.append(line)
                
                if track_updates:
                    update_tracking.append({
                        'file': file_path,
                        'folder': os.path.dirname(file_path),
                        'parameter': current_param,
                        'old_value': old_block.rstrip(),
                        'new_value': '',
                        'status': 'Updated',
                        '_building_new': new_block_lines
                    })
            
            elif stripped_line.startswith('TEXT') and current_param and current_param in config_parameters:
                old_block = ''.join(current_block_lines)
                new_value = config_parameters[current_param]
                indentation = line[:len(line) - len(line.lstrip())]
                line = f"{indentation}TEXT {new_value}\\n"
                file_was_modified = True
                
                new_block_lines = []
                for bl in current_block_lines[:-1]:
                    new_block_lines.append(bl)
                new_block_lines.append(line)
                
                if track_updates:
                    update_tracking.append({
                        'file': file_path,
                        'folder': os.path.dirname(file_path),
                        'parameter': current_param,
                        'old_value': old_block.rstrip(),
                        'new_value': '',
                        'status': 'Updated',
                        '_building_new': new_block_lines
                    })
            
            updated_lines.append(line)
            
            if stripped_line == 'END' and track_updates and update_tracking:
                for track_entry in reversed(update_tracking):
                    if '_building_new' in track_entry and not track_entry['new_value']:
                        track_entry['_building_new'].append(line)
                        track_entry['new_value'] = ''.join(track_entry['_building_new']).rstrip()
                        del track_entry['_building_new']
                        break
                current_param = None
                current_block_lines = []
        
        if file_was_modified:
            try:
                with open(file_path, 'w', encoding=encoding_used) as file:
                    file.writelines(updated_lines)
                updated_files.append(file_path)
            except Exception as e:
                print(f"Error writing to file {os.path.basename(file_path)}: {e}")
    
    if track_updates:
        return updated_files, update_tracking
    return updated_files

def update_specific_files(file_list, config_parameters, track_updates=False):
    """Update specific files directly (avoiding directory scan)."""
    updated_files = []
    update_tracking = []
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    for file_path in file_list:
        try:
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    lines = file.readlines()
                encoding_used = 'utf-8'
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='latin-1') as file:
                    lines = file.readlines()
                encoding_used = 'latin-1'
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            continue

        updated_lines = []
        current_param = None
        current_block_lines = []
        in_block = False
        file_was_modified = False
        
        for i, line in enumerate(lines):
            stripped_line = line.strip()
            
            if stripped_line.startswith(keywords):
                parts = stripped_line.split()
                if len(parts) > 1:
                    current_param = parts[1]
                    in_block = True
                    current_block_lines = [line]
                    updated_lines.append(line)
                    continue
            
            if in_block:
                current_block_lines.append(line)
                
                if stripped_line == 'END':
                    in_block = False
                    if current_param not in config_parameters:
                        current_param = None
                        current_block_lines = []
            
            if stripped_line.startswith('WERT') and current_param and current_param in config_parameters:
                # Capture old block before modification
                old_block = ''.join(current_block_lines)
                
                new_value = config_parameters[current_param]
                indentation = line[:len(line) - len(line.lstrip())]
                line = f"{indentation}WERT {new_value}\n"
                file_was_modified = True
                
                # Build new block
                new_block_lines = []
                for bl in current_block_lines[:-1]:
                    new_block_lines.append(bl)
                new_block_lines.append(line)
                
                if track_updates:
                    update_tracking.append({
                        'file': file_path,
                        'folder': os.path.dirname(file_path),
                        'parameter': current_param,
                        'old_value': old_block.rstrip(),
                        'new_value': '',
                        'status': 'Updated',
                        '_building_new': new_block_lines
                    })
            
            elif stripped_line.startswith('TEXT') and current_param and current_param in config_parameters:
                # Capture old block before modification
                old_block = ''.join(current_block_lines)
                
                new_value = config_parameters[current_param]
                indentation = line[:len(line) - len(line.lstrip())]
                line = f"{indentation}TEXT {new_value}\n"
                file_was_modified = True
                
                # Build new block
                new_block_lines = []
                for bl in current_block_lines[:-1]:
                    new_block_lines.append(bl)
                new_block_lines.append(line)
                
                if track_updates:
                    update_tracking.append({
                        'file': file_path,
                        'folder': os.path.dirname(file_path),
                        'parameter': current_param,
                        'old_value': old_block.rstrip(),
                        'new_value': '',
                        'status': 'Updated',
                        '_building_new': new_block_lines
                    })
            
            updated_lines.append(line)
            
            # When we hit END, complete the new_value block in tracking
            if stripped_line == 'END' and track_updates and update_tracking:
                for track_entry in reversed(update_tracking):
                    if '_building_new' in track_entry and not track_entry['new_value']:
                        track_entry['_building_new'].append(line)
                        track_entry['new_value'] = ''.join(track_entry['_building_new']).rstrip()
                        del track_entry['_building_new']
                        break
                current_param = None
                current_block_lines = []
        
        if file_was_modified:
            try:
                with open(file_path, 'w', encoding=encoding_used) as file:
                    file.writelines(updated_lines)
                updated_files.append(file_path)
            except Exception as e:
                print(f"Error writing to file {os.path.basename(file_path)}: {e}")
    
    if track_updates:
        return updated_files, update_tracking
    return updated_files

def remove_parameters_from_files_optimized(file_cache_dict, parameters_to_remove, track_removals=False):
    """Remove entire parameter blocks from files using cached content."""
    updated_files = []
    removal_tracking = []
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    for file_path, cache_data in file_cache_dict.items():
        content = cache_data['content']
        encoding_used = cache_data['encoding']
        lines = content.split('\\n')
        
        # Convert back to line format with newlines
        lines = [line + '\\n' for line in lines]
        if lines and not lines[-1].endswith('\\n'):
            lines[-1] = lines[-1].rstrip('\\n')

        updated_lines = []
        current_param = None
        in_block = False
        skip_block = False
        removed_block = []
        file_was_modified = False
        
        for i, line in enumerate(lines):
            stripped_line = line.strip()
            
            if stripped_line.startswith(keywords):
                parts = stripped_line.split()
                if len(parts) > 1:
                    current_param = parts[1]
                    in_block = True
                    
                    if current_param in parameters_to_remove:
                        skip_block = True
                        removed_block = [line]
                        file_was_modified = True
                        continue
                    else:
                        skip_block = False
            
            if in_block and skip_block:
                removed_block.append(line)
                
                if stripped_line == 'END':
                    in_block = False
                    skip_block = False
                    
                    if track_removals:
                        removal_tracking.append({
                            'file': file_path,
                            'folder': os.path.dirname(file_path),
                            'parameter': current_param,
                            'removed_block': ''.join(removed_block).rstrip(),
                            'status': 'Removed'
                        })
                    
                    removed_block = []
                    current_param = None
                continue
            
            if in_block and stripped_line == 'END':
                in_block = False
                current_param = None
            
            if not skip_block:
                updated_lines.append(line)
        
        if file_was_modified:
            try:
                with open(file_path, 'w', encoding=encoding_used) as file:
                    file.writelines(updated_lines)
                updated_files.append(file_path)
            except Exception as e:
                print(f"Error writing to file {os.path.basename(file_path)}: {e}")
    
    if track_removals:
        return updated_files, removal_tracking
    return updated_files

def remove_parameters_from_files(file_list, parameters_to_remove, track_removals=False):
    """Remove entire parameter blocks from files."""
    updated_files = []
    removal_tracking = []
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    for file_path in file_list:
        try:
            try:
                with open(file_path, 'r', encoding='utf-8') as file:
                    lines = file.readlines()
                encoding_used = 'utf-8'
            except UnicodeDecodeError:
                with open(file_path, 'r', encoding='latin-1') as file:
                    lines = file.readlines()
                encoding_used = 'latin-1'
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            continue

        updated_lines = []
        current_param = None
        in_block = False
        skip_block = False
        removed_block = []
        file_was_modified = False
        
        for i, line in enumerate(lines):
            stripped_line = line.strip()
            
            # Start of a parameter block
            if stripped_line.startswith(keywords):
                parts = stripped_line.split()
                if len(parts) > 1:
                    current_param = parts[1]
                    in_block = True
                    
                    # Check if this parameter should be removed
                    if current_param in parameters_to_remove:
                        skip_block = True
                        removed_block = [line]
                        file_was_modified = True
                        continue
                    else:
                        skip_block = False
            
            # Inside a block that should be removed
            if in_block and skip_block:
                removed_block.append(line)
                
                # End of block to remove
                if stripped_line == 'END':
                    in_block = False
                    skip_block = False
                    
                    # Track removal
                    if track_removals:
                        removal_tracking.append({
                            'file': file_path,
                            'folder': os.path.dirname(file_path),
                            'parameter': current_param,
                            'removed_block': ''.join(removed_block).rstrip(),
                            'status': 'Removed'
                        })
                    
                    removed_block = []
                    current_param = None
                continue
            
            # End of normal block
            if in_block and stripped_line == 'END':
                in_block = False
                current_param = None
            
            # Keep lines that are not in removed blocks
            if not skip_block:
                updated_lines.append(line)
        
        # Write file if modified
        if file_was_modified:
            try:
                with open(file_path, 'w', encoding=encoding_used) as file:
                    file.writelines(updated_lines)
                updated_files.append(file_path)
            except Exception as e:
                print(f"Error writing to file {os.path.basename(file_path)}: {e}")
    
    if track_removals:
        return updated_files, removal_tracking
    return updated_files

def process_multiple_directories(directories, config_parameters, file_extension=".dcm"):
    """Process multiple directories (tuần tự)."""
    all_matching_files = []
    all_updated_files = []
    
    for directory in directories:
        matching_files = find_files_with_parameters(directory, config_parameters, file_extension)
        all_matching_files.extend(matching_files)
        updated_files = update_files_in_directory(directory, config_parameters, file_extension)
        all_updated_files.extend(updated_files)
    
    return all_matching_files, all_updated_files

class MultiDirectoryDialog:
    def __init__(self, parent):
        self.parent = parent
        self.selected_directories = []
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Multiple Directories")
        self.dialog.geometry("700x800")
        self.dialog.grab_set()
        
        # Load parent folder history
        self.config_file = os.path.join(os.path.dirname(__file__), 'dcm_parameter_config.json')
        self.parent_folder_history = []
        self.load_parent_history()
        
        self.setup_dialog()
        
    def setup_dialog(self):
        main_frame = tk.Frame(self.dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        instruction_text = (
            "Select multiple directories:\n"
            "• Use 'Browse Parent Folder' to navigate to a parent directory\n"
            "• Use Ctrl+A to select all folders\n"
            "• Use Ctrl+Click or Shift+Click for multiple selection"
        )
        tk.Label(main_frame, text=instruction_text, font=('Arial', 10),
                 justify=tk.LEFT, fg="blue").pack(anchor='w', pady=(0, 15))
        
        path_frame = tk.Frame(main_frame)
        path_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Label(path_frame, text="Parent Folder:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
        
        # Combobox for parent folder with history
        path_combo_frame = tk.Frame(path_frame)
        path_combo_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 10))
        
        self.current_path_var = tk.StringVar(value="")
        self.path_combo = ttk.Combobox(path_combo_frame, textvariable=self.current_path_var, 
                                       font=('Arial', 9))
        self.path_combo['values'] = self.parent_folder_history
        if self.parent_folder_history:
            self.current_path_var.set(self.parent_folder_history[0])
        self.path_combo.pack(fill=tk.X)
        self.path_combo.bind('<<ComboboxSelected>>', self.on_path_selected)
        
        # Browse button
        tk.Button(path_frame, text="Browse", command=self.browse_parent_folder,
                  width=10, bg='lightgreen').pack(side=tk.LEFT)
        
        tk.Label(main_frame, text="Available Folders in Parent Directory:",
                 font=('Arial', 10, 'bold')).pack(anchor='w', pady=(10, 5))
        
        available_frame = tk.Frame(main_frame)
        available_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.available_listbox = tk.Listbox(available_frame, selectmode=tk.EXTENDED, height=8)
        available_scrollbar = tk.Scrollbar(available_frame, orient=tk.VERTICAL,
                                           command=self.available_listbox.yview)
        self.available_listbox.configure(yscrollcommand=available_scrollbar.set)
        self.available_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        available_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.available_listbox.bind('<Control-a>', self.select_all_available)
        self.available_listbox.bind('<Return>', self.add_selected_directories)
        self.available_listbox.bind('<Double-Button-1>', self.add_selected_directories)
        
        # Load subdirectories AFTER listbox is created
        if self.parent_folder_history:
            self.load_subdirectories(self.parent_folder_history[0])
        
        transfer_frame = tk.Frame(main_frame)
        transfer_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Button(transfer_frame, text="Add Selected Folders →", command=self.add_selected_directories,
                  width=20, bg='lightblue').pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(transfer_frame, text="Add All Folders →", command=self.add_all_directories,
                  width=15).pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Label(main_frame, text="Selected Directories:", font=('Arial', 10, 'bold')).pack(anchor='w', pady=(10, 5))
        
        selected_frame = tk.Frame(main_frame)
        selected_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.selected_listbox = tk.Listbox(selected_frame, selectmode=tk.EXTENDED, height=6)
        selected_scrollbar = tk.Scrollbar(selected_frame, orient=tk.VERTICAL,
                                          command=self.selected_listbox.yview)
        self.selected_listbox.configure(yscrollcommand=selected_scrollbar.set)
        self.selected_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        selected_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.selected_listbox.bind('<Control-a>', self.select_all_selected)
        self.selected_listbox.bind('<Delete>', self.remove_selected_directories)
        
        mgmt_frame = tk.Frame(main_frame)
        mgmt_frame.pack(fill=tk.X, pady=(0, 10))
        tk.Button(mgmt_frame, text="Remove Selected", command=self.remove_selected_directories,
                  width=15).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(mgmt_frame, text="Clear All", command=self.clear_all, width=10).pack(side=tk.LEFT)
        
        self.count_label = tk.Label(mgmt_frame, text="Selected: 0 directories", fg="blue",
                                    font=('Arial', 10, 'bold'))
        self.count_label.pack(side=tk.RIGHT)
        
        dialog_button_frame = tk.Frame(main_frame)
        dialog_button_frame.pack(fill=tk.X, pady=(10, 0))
        tk.Button(dialog_button_frame, text="OK", command=self.ok_clicked,
                  width=10, bg='lightgreen', font=('Arial', 10, 'bold')).pack(side=tk.RIGHT, padx=(5, 0))
        tk.Button(dialog_button_frame, text="Cancel", command=self.cancel_clicked,
                  width=10, font=('Arial', 10)).pack(side=tk.RIGHT)
        
        self.center_dialog()
        
    def center_dialog(self):
        self.dialog.transient(self.parent)
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
    
    def load_parent_history(self):
        """Load parent folder history from config file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.parent_folder_history = config.get('parent_folder_history', [])
        except Exception as e:
            print(f"Error loading parent folder history: {e}")
            self.parent_folder_history = []
    
    def save_parent_history(self):
        """Save parent folder history to config file"""
        try:
            config = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            config['parent_folder_history'] = self.parent_folder_history
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving parent folder history: {e}")
    
    def on_path_selected(self, event=None):
        """Handle selection from dropdown"""
        selected_path = self.current_path_var.get()
        if selected_path and os.path.exists(selected_path):
            self.load_subdirectories(selected_path)
        
    def browse_parent_folder(self):
        parent_dir = filedialog.askdirectory(title="Select Parent Directory")
        if parent_dir:
            self.current_path_var.set(parent_dir)
            self.load_subdirectories(parent_dir)
            
            # Save to history
            if parent_dir not in self.parent_folder_history:
                self.parent_folder_history.insert(0, parent_dir)
                self.parent_folder_history = self.parent_folder_history[:10]  # Keep last 10
                self.path_combo['values'] = self.parent_folder_history
                self.save_parent_history()
            
    def load_subdirectories(self, parent_dir):
        self.available_listbox.delete(0, tk.END)
        try:
            directories = []
            # Use os.walk to recursively find all subdirectories
            for root, dirs, files in os.walk(parent_dir):
                for dir_name in dirs:
                    dir_path = os.path.join(root, dir_name)
                    directories.append(dir_path)
            directories.sort()
            for directory in directories:
                self.available_listbox.insert(tk.END, directory)
        except Exception as e:
            messagebox.showerror("Error", f"Error reading directory: {str(e)}")
    
    def select_all_available(self, event=None):
        self.available_listbox.select_set(0, tk.END)
        return 'break'
    
    def select_all_selected(self, event=None):
        self.selected_listbox.select_set(0, tk.END)
        return 'break'
        
    def add_selected_directories(self, event=None):
        selections = self.available_listbox.curselection()
        for index in selections:
            directory = self.available_listbox.get(index)
            if directory not in self.selected_directories:
                self.selected_directories.append(directory)
                self.selected_listbox.insert(tk.END, directory)
        self.update_count()
        
    def add_all_directories(self):
        for i in range(self.available_listbox.size()):
            directory = self.available_listbox.get(i)
            if directory not in self.selected_directories:
                self.selected_directories.append(directory)
                self.selected_listbox.insert(tk.END, directory)
        self.update_count()
        
    def remove_selected_directories(self, event=None):
        selections = self.selected_listbox.curselection()
        for index in reversed(selections):
            directory = self.selected_listbox.get(index)
            self.selected_listbox.delete(index)
            self.selected_directories.remove(directory)
        self.update_count()
        
    def clear_all(self):
        self.selected_listbox.delete(0, tk.END)
        self.selected_directories.clear()
        self.update_count()
        
    def update_count(self):
        count = len(self.selected_directories)
        self.count_label.config(text=f"Selected: {count} directories")
        
    def ok_clicked(self):
        self.dialog.destroy()
        
    def cancel_clicked(self):
        self.selected_directories.clear()
        self.dialog.destroy()

class SingleFolderDialog:
    """Dialog for selecting a single folder with dropdown history."""
    def __init__(self, parent, folder_history, config_file):
        self.parent = parent
        self.folder_history = folder_history
        self.config_file = config_file
        self.selected_folder = None
        
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Single Folder")
        self.dialog.geometry("600x200")
        self.dialog.grab_set()
        
        self.setup_dialog()
        
        # Center window
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
        
        self.dialog.wait_window()
    
    def setup_dialog(self):
        main_frame = tk.Frame(self.dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Title
        tk.Label(main_frame, text="Select a folder:", font=('Segoe UI', 10, 'bold')).pack(anchor='w', pady=(0, 10))
        
        # Folder selection frame
        folder_frame = tk.Frame(main_frame)
        folder_frame.pack(fill=tk.X, pady=(0, 20))
        
        tk.Label(folder_frame, text="Folder:", font=('Segoe UI', 9)).pack(side=tk.LEFT, padx=(0, 10))
        
        # Combobox with history
        self.folder_var = tk.StringVar()
        self.folder_combo = ttk.Combobox(folder_frame, textvariable=self.folder_var, 
                                         font=('Segoe UI', 9), width=50)
        self.folder_combo['values'] = self.folder_history
        if self.folder_history:
            self.folder_var.set(self.folder_history[0])
        self.folder_combo.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        
        # Browse button
        tk.Button(folder_frame, text="Browse", command=self.browse_folder,
                 bg='#2196F3', fg='white', font=('Segoe UI', 8, 'bold'),
                 relief=tk.FLAT, cursor='hand2', width=10).pack(side=tk.LEFT)
        
        # Action buttons
        button_frame = tk.Frame(main_frame)
        button_frame.pack()
        
        tk.Button(button_frame, text="OK", command=self.ok_clicked,
                 bg='#4CAF50', fg='white', font=('Segoe UI', 9, 'bold'),
                 relief=tk.FLAT, cursor='hand2', width=12, pady=6).pack(side=tk.LEFT, padx=(0, 10))
        tk.Button(button_frame, text="Cancel", command=self.cancel_clicked,
                 bg='#F44336', fg='white', font=('Segoe UI', 9, 'bold'),
                 relief=tk.FLAT, cursor='hand2', width=12, pady=6).pack(side=tk.LEFT)
    
    def browse_folder(self):
        folder = filedialog.askdirectory(title="Select Directory")
        if folder:
            self.folder_var.set(folder)
    
    def ok_clicked(self):
        folder = self.folder_var.get()
        if folder and os.path.exists(folder):
            self.selected_folder = folder
        self.dialog.destroy()
    
    def cancel_clicked(self):
        self.selected_folder = None
        self.dialog.destroy()
    
    def get_result(self):
        return self.selected_folder

class ParameterUpdateTool:
    def __init__(self, root):
        self.root = root
        self.root.title("DCM File Parameter Update Tool - Multi Directory")
        self.root.geometry("800x500")
        
        # Modern color scheme
        self.colors = {
            'bg': '#f0f0f0',
            'panel': '#ffffff',
            'primary': '#2196F3',
            'success': '#4CAF50',
            'warning': '#FF9800',
            'danger': '#F44336',
            'text': '#333333',
            'border': '#ddd'
        }
        
        self.root.configure(bg=self.colors['bg'])
        
        # Config file path
        self.config_file = os.path.join(os.path.dirname(__file__), 'dcm_parameter_config.json')
        
        self.selected_directories = []
        self.config_file_history = []
        self.single_folder_history = []  # Initialize single folder history
        self._busy = False
        self._lock = threading.Lock()
        
        # Excel report tracking
        self.update_history = []  # Track all parameter updates for Excel report
        self.removal_history = []  # Track all parameter removals
        self.config_parameters = {}  # Store config parameters
        self.file_contents_cache = {}  # Cache file contents for Excel generation
        self.matching_files_cache = []  # Store list of files that need updating (avoid re-scanning)
        
        self.setup_gui()
        self.load_config()
        
        # Save config on window close
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

    def _set_busy(self, value: bool):
        with self._lock:
            self._busy = value
        self.root.configure(cursor="watch" if value else "")
        self.root.update_idletasks()

    def setup_gui(self):
        main_frame = tk.Frame(self.root, bg=self.colors['bg'], padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure([1, 3, 6], weight=0)
        main_frame.rowconfigure(8, weight=3)

        tk.Label(main_frame, text="Select Parameter Config File:", font=('Segoe UI', 10, 'bold'),
                bg=self.colors['bg'], fg=self.colors['text']).grid(row=0, column=0, sticky='w', pady=(0, 5))
        file_frame = tk.Frame(main_frame, bg=self.colors['bg'])
        file_frame.grid(row=1, column=0, columnspan=3, sticky='ew', pady=(0, 15))
        file_frame.columnconfigure(0, weight=1)
        self.list_file_combo = ttk.Combobox(file_frame, font=('Segoe UI', 9))
        self.list_file_combo.grid(row=0, column=0, sticky='ew', padx=(0, 10))
        tk.Button(file_frame, text="Browse", command=self.open_file,
                 bg=self.colors['primary'], fg='white', font=('Segoe UI', 9, 'bold'),
                 relief=tk.FLAT, padx=15, pady=6, cursor='hand2').grid(row=0, column=1)

        tk.Label(main_frame, text="Select Multiple Directories containing DCM files:", font=('Segoe UI', 10, 'bold'),
                bg=self.colors['bg'], fg=self.colors['text']).grid(row=2, column=0, sticky='w', pady=(0, 5))
        dir_frame = tk.Frame(main_frame, bg=self.colors['bg'])
        dir_frame.grid(row=3, column=0, columnspan=3, sticky='ew', pady=(0, 15))
        dir_frame.columnconfigure(0, weight=1)

        listbox_frame = tk.Frame(dir_frame, bg=self.colors['panel'], relief=tk.SOLID, borderwidth=1)
        listbox_frame.grid(row=0, column=0, sticky='ew', padx=(0, 10))
        listbox_frame.columnconfigure(0, weight=1)
        self.directory_listbox = tk.Listbox(listbox_frame, height=4, font=('Segoe UI', 9),
                                           relief=tk.FLAT, bg=self.colors['panel'])
        self.directory_listbox.grid(row=0, column=0, sticky='ew')
        listbox_scrollbar = tk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.directory_listbox.yview)
        listbox_scrollbar.grid(row=0, column=1, sticky='ns')
        self.directory_listbox.configure(yscrollcommand=listbox_scrollbar.set)

        dir_buttons_frame = tk.Frame(dir_frame, bg=self.colors['bg'])
        dir_buttons_frame.grid(row=0, column=1, sticky='n')
        tk.Button(dir_buttons_frame, text="Select Folders", command=self.select_multiple_directories,
                  width=14, bg=self.colors['primary'], fg='white', font=('Segoe UI', 8, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=4).pack(pady=(0, 5))
        tk.Button(dir_buttons_frame, text="Add Single Folder", command=self.add_single_directory,
                  width=14, bg=self.colors['success'], fg='white', font=('Segoe UI', 8, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=4).pack(pady=(0, 5))
        tk.Button(dir_buttons_frame, text="Remove Selected", command=self.remove_directory,
                  width=14, bg=self.colors['warning'], fg='white', font=('Segoe UI', 8, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=4).pack(pady=(0, 5))
        tk.Button(dir_buttons_frame, text="Clear All", command=self.clear_directories,
                  width=14, bg=self.colors['danger'], fg='white', font=('Segoe UI', 8, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=4).pack()

        ext_frame = tk.Frame(main_frame, bg=self.colors['bg'])
        ext_frame.grid(row=4, column=0, columnspan=3, sticky='ew', pady=(0, 15))
        tk.Label(ext_frame, text="File Extension:", bg=self.colors['bg'],
                fg=self.colors['text'], font=('Segoe UI', 9)).grid(row=0, column=0, sticky='w')
        self.extension_var = tk.StringVar(value=".dcm")
        ext_combo = ttk.Combobox(ext_frame, textvariable=self.extension_var,
                                values=[".dcm", ".cfg", ".txt"], width=10, font=('Segoe UI', 9))
        ext_combo.grid(row=0, column=1, padx=(10, 0), sticky='w')
        self.dir_count_label = tk.Label(ext_frame, text="Selected directories: 0",
                                       bg=self.colors['bg'], fg=self.colors['primary'],
                                       font=('Segoe UI', 9, 'bold'))
        self.dir_count_label.grid(row=0, column=2, padx=(20, 0), sticky='w')

        button_frame = tk.Frame(main_frame, bg=self.colors['bg'])
        button_frame.grid(row=5, column=0, columnspan=3, pady=(10, 0))
        tk.Button(button_frame, text="Preview Changes", command=self.preview_changes,
                  bg=self.colors['primary'], fg='white', width=14, font=('Segoe UI', 9, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=6).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(button_frame, text="Update Files", command=self.start_update,
                  bg=self.colors['success'], fg='white', width=14, font=('Segoe UI', 9, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=6).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(button_frame, text="Remove Params", command=self.start_remove,
                  bg=self.colors['danger'], fg='white', width=14, font=('Segoe UI', 9, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=6).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(button_frame, text="📊 Export Excel", command=self.export_excel_report,
                  bg='#9C27B0', fg='white', width=14, font=('Segoe UI', 9, 'bold'),
                  relief=tk.FLAT, cursor='hand2', pady=6).pack(side=tk.LEFT)

        tk.Label(main_frame, text="Results:", font=('Segoe UI', 10, 'bold'),
                bg=self.colors['bg'], fg=self.colors['text']).grid(row=6, column=0, sticky='w', pady=(20, 5))
        text_frame = tk.Frame(main_frame, bg=self.colors['panel'], relief=tk.SOLID, borderwidth=1)
        text_frame.grid(row=7, column=0, columnspan=3, sticky='nsew', pady=(0, 10))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        self.results_text = tk.Text(text_frame, height=12, wrap=tk.WORD, font=('Segoe UI', 9),
                                   relief=tk.FLAT, bg=self.colors['panel'])
        scrollbar = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        self.results_text.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        
        # Configure text tags for highlighting
        self.results_text.tag_configure('diff_highlight', background='#FFFF99', foreground='#D32F2F', font=('Segoe UI', 9, 'bold'))
        self.results_text.tag_configure('param_name', foreground='#1976D2', font=('Segoe UI', 9, 'bold'))
        self.results_text.tag_configure('success', foreground='#388E3C', font=('Segoe UI', 9, 'bold'))
        self.results_text.tag_configure('warning', foreground='#F57C00', font=('Segoe UI', 9, 'bold'))
        self.results_text.tag_configure('danger', foreground='#C00000', font=('Segoe UI', 9, 'bold'))
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(7, weight=1)

    def load_config(self):
        """Load saved configuration from JSON file."""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    
                # Load config file history
                if 'config_file_history' in config:
                    self.config_file_history = config['config_file_history']
                    self.list_file_combo['values'] = self.config_file_history
                    if self.config_file_history:
                        self.list_file_combo.set(self.config_file_history[0])
                
                # Load single folder history
                if 'single_folder_history' in config:
                    self.single_folder_history = config['single_folder_history']
                
                # Load directories
                if 'directories' in config:
                    for directory in config['directories']:
                        if os.path.isdir(directory) and directory not in self.selected_directories:
                            self.selected_directories.append(directory)
                            self.directory_listbox.insert(tk.END, directory)
                    self.update_directory_count()
                    
                self.log_message(f"Loaded configuration: {len(self.selected_directories)} directories")
        except Exception as e:
            self.log_message(f"Could not load config: {e}")
    
    def save_config(self):
        """Save current configuration to JSON file."""
        try:
            # Update history list
            config_file_path = self.list_file_combo.get()
            if config_file_path and config_file_path not in self.config_file_history:
                self.config_file_history.insert(0, config_file_path)
                self.config_file_history = self.config_file_history[:10]  # Keep last 10
            
            config = {
                'config_file_history': self.config_file_history,
                'single_folder_history': self.single_folder_history,
                'directories': self.selected_directories
            }
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
                
        except Exception as e:
            print(f"Could not save config: {e}")
    
    def on_closing(self):
        """Handle window closing event."""
        self.save_config()
        self.root.destroy()

    def log_message(self, message, tag=None):
        timestamp = f"{datetime.now().strftime('%H:%M:%S')} - "
        self.results_text.insert(tk.END, timestamp)
        self.results_text.insert(tk.END, f"{message}\n", tag)
        self.results_text.see(tk.END)
    
    def log_parameter_diff(self, param_name, current_value, target_value, file_name):
        """Log parameter with highlight if values differ"""
        timestamp = f"{datetime.now().strftime('%H:%M:%S')} - "
        self.results_text.insert(tk.END, timestamp)
        self.results_text.insert(tk.END, f"  Parameter: ", None)
        self.results_text.insert(tk.END, f"{param_name}", 'param_name')
        
        if current_value != target_value:
            self.results_text.insert(tk.END, f" [NEEDS UPDATE] ", 'diff_highlight')
            self.results_text.insert(tk.END, f"Current: {current_value} → Target: {target_value}\n", None)
        else:
            self.results_text.insert(tk.END, f" [OK] Current: {current_value}\n", 'success')
        
        self.results_text.see(tk.END)

    def open_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Parameter Config File",
            filetypes=[("Text files", "*.txt"), ("Config files", "*.cfg"), ("All files", "*.*")]
        )
        if file_path:
            self.list_file_combo.set(file_path)
            # Add to history
            if file_path not in self.config_file_history:
                self.config_file_history.insert(0, file_path)
                self.config_file_history = self.config_file_history[:10]
                self.list_file_combo['values'] = self.config_file_history
            self.save_config()

    def select_multiple_directories(self):
        dialog = MultiDirectoryDialog(self.root)
        self.root.wait_window(dialog.dialog)
        for directory in dialog.selected_directories:
            if directory not in self.selected_directories:
                self.selected_directories.append(directory)
                self.directory_listbox.insert(tk.END, directory)
        self.update_directory_count()
        self.save_config()

    def add_single_directory(self):
        # Show dialog with dropdown history
        dialog = SingleFolderDialog(self.root, self.single_folder_history, self.config_file)
        directory_path = dialog.get_result()
        
        if directory_path and directory_path not in self.selected_directories:
            self.selected_directories.append(directory_path)
            self.directory_listbox.insert(tk.END, directory_path)
            self.update_directory_count()
            
            # Update history
            if directory_path not in self.single_folder_history:
                self.single_folder_history.insert(0, directory_path)
                self.single_folder_history = self.single_folder_history[:10]
            
            self.save_config()

    def remove_directory(self):
        selection = self.directory_listbox.curselection()
        if selection:
            index = selection[0]
            self.directory_listbox.delete(index)
            del self.selected_directories[index]
            self.update_directory_count()
            self.save_config()

    def clear_directories(self):
        self.directory_listbox.delete(0, tk.END)
        self.selected_directories.clear()
        self.update_directory_count()
        self.save_config()

    def update_directory_count(self):
        count = len(self.selected_directories)
        self.dir_count_label.config(text=f"Selected directories: {count}")

    def validate_inputs(self):
        config_path = self.list_file_combo.get().strip()
        if not config_path:
            messagebox.showerror("Error", "Please select a config file.")
            return False
        if not self.selected_directories:
            messagebox.showerror("Error", "Please select at least one directory.")
            return False
        if not os.path.isfile(config_path):
            messagebox.showerror("Error", "Please select a valid config file.")
            return False
        invalid_dirs = [d for d in self.selected_directories if not os.path.isdir(d)]
        if invalid_dirs:
            messagebox.showerror("Error", f"Invalid directories found:\n{chr(10).join(invalid_dirs)}")
            return False
        return True

    def preview_changes(self):
        if self._busy:
            messagebox.showinfo("Info", "Task đang chạy, vui lòng đợi.")
            return
        if not self.validate_inputs():
            return
        self.results_text.delete(1.0, tk.END)
        self.log_message("Starting preview (multi-thread)...")
        self._set_busy(True)

        def worker():
            logs = []
            try:
                config_path = self.list_file_combo.get().strip()
                ext = self.extension_var.get()
                config_parameters = read_config_file(config_path)
                logs.append(f"Loaded {len(config_parameters)} parameters.")
                for p, v in config_parameters.items():
                    logs.append(f"  {p} -> {v}")
                
                # Store config parameters for later use
                self.config_parameters = config_parameters

                logs.append(f"\nScanning {len(self.selected_directories)} directories in parallel...")
                all_matching = []

                def find_only(d):
                    return d, find_files_with_parameters(d, config_parameters, ext)

                workers = min(8, len(self.selected_directories) or 1)
                with ThreadPoolExecutor(max_workers=workers) as ex:
                    futures = {ex.submit(find_only, d): d for d in self.selected_directories}
                    for fut in as_completed(futures):
                        d, found = fut.result()
                        logs.append(f"\nDirectory: {d}")
                        if found:
                            logs.append(f"  Found {len(found)} matching files:")
                            for fp in found:
                                logs.append(f"    - {os.path.basename(fp)}")
                                # Analyze parameters in the file
                                file_params = extract_all_parameters_from_file(fp)
                                for param in file_params:
                                    if param['name'] in config_parameters:
                                        current_val = param['value']
                                        target_val = config_parameters[param['name']]
                                        logs.append(('param_diff', param['name'], current_val, target_val, os.path.basename(fp)))
                            all_matching.extend(found)
                        else:
                            logs.append("  No matching files")

                if not all_matching:
                    logs.append("\nNo files found with the specified parameters.")
                else:
                    logs.append(f"\nTotal: {len(all_matching)} files would be updated.")
                    
                    # Cache matching files list and original file contents during preview
                    logs.append("\nCaching original file contents...")
                    self.matching_files_cache = all_matching  # Store matching files
                    self.file_contents_cache = {}
                    for file_path in all_matching:
                        self.file_contents_cache[file_path] = extract_all_parameters_from_file(file_path)
                    logs.append(f"Cached {len(self.file_contents_cache)} files with original values")
            except Exception as e:
                logs.append(f"Error during preview: {e}")
                import traceback
                logs.append(traceback.format_exc())
            finally:
                self.root.after(0, lambda: self._apply_preview_logs(logs))

        threading.Thread(target=worker, daemon=True).start()

    def _apply_preview_logs(self, logs):
        for item in logs:
            if isinstance(item, tuple) and item[0] == 'param_diff':
                # Special handling for parameter diff logs
                _, param_name, current_val, target_val, file_name = item
                self.log_parameter_diff(param_name, current_val, target_val, file_name)
            else:
                self.log_message(item)
        self._set_busy(False)

    def start_update(self):
        if self._busy:
            messagebox.showinfo("Info", "Task đang chạy, vui lòng đợi.")
            return
        if not self.validate_inputs():
            return
        result = messagebox.askyesno(
            "Confirm Update",
            f"Update files in {len(self.selected_directories)} directories (multi-thread)?"
        )
        if not result:
            return

        self.results_text.delete(1.0, tk.END)
        self.log_message("Starting file update (multi-thread)...")
        self._set_busy(True)

        def worker():
            logs = []
            try:
                ext = self.extension_var.get()
                config_path = self.list_file_combo.get().strip()
                config_parameters = read_config_file(config_path)
                if not config_parameters:
                    logs.append("No valid parameters found in config file.")
                    self.root.after(0, lambda: self._finish_update(logs, [], []))
                    return

                logs.append(f"Loaded {len(config_parameters)} parameters:")
                for p, v in config_parameters.items():
                    logs.append(f"  {p} -> {v}")
                
                # Store config parameters for Excel export
                self.config_parameters = config_parameters

                # Check if we already have matching files from preview
                if self.matching_files_cache and self.file_contents_cache:
                    logs.append(f"\nUsing cached file list from preview ({len(self.matching_files_cache)} files)")
                    logs.append("Skipping file scan - updating cached files directly...")
                    matching_files = self.matching_files_cache
                else:
                    logs.append(f"\nNo cached files found. Scanning files in {len(self.selected_directories)} directories...")
                    # Cache original file contents BEFORE updating
                    matching_files = []
                    self.file_contents_cache = {}
                    for directory in self.selected_directories:
                        found_files = find_files_with_parameters(directory, config_parameters, ext)
                        matching_files.extend(found_files)
                        for file_path in found_files:
                            self.file_contents_cache[file_path] = extract_all_parameters_from_file(file_path)
                    logs.append(f"Cached {len(self.file_contents_cache)} files with original values")
                
                if not matching_files:
                    logs.append("\nNo files found with the specified parameters.")
                    self.root.after(0, lambda: self._finish_update(logs, [], []))
                    return

                logs.append(f"\nUpdating {len(matching_files)} files in parallel...")
                
                # Update files directly with tracking
                all_updated = []
                all_tracking = []
                
                # Group files by directory for organized updates
                files_by_dir = {} 
                for file_path in matching_files:
                    dir_path = os.path.dirname(file_path)
                    if dir_path not in files_by_dir:
                        files_by_dir[dir_path] = []
                    files_by_dir[dir_path].append(file_path)
                
                # Update each directory's files
                for dir_path, file_list in files_by_dir.items():
                    updated, tracking = update_specific_files(file_list, config_parameters, track_updates=True)
                    all_updated.extend(updated)
                    all_tracking.extend(tracking)
                
                # Store update tracking data
                self.update_history = all_tracking

                logs.append(f"\nProcessed {len(matching_files)} files containing target parameters.")
                if all_updated:
                    logs.append(f"\n✓ Successfully updated {len(all_updated)} files:")
                    grouped = {}
                    for fp in all_updated:
                        grouped.setdefault(os.path.dirname(fp), []).append(fp)
                    for d, file_paths in grouped.items():
                        logs.append(f"\nDirectory: {d}")
                        for fp in file_paths:
                            logs.append(f"  - {os.path.basename(fp)}")
                            # Show what was changed in this file
                            for track in all_tracking:
                                if track['file'] == fp:
                                    # Extract old and new values from the tracking data
                                    param_name = track['parameter']
                                    # Parse old value from old_value block
                                    old_val = ''
                                    for line in track['old_value'].split('\n'):
                                        if line.strip().startswith('WERT') or line.strip().startswith('TEXT'):
                                            old_val = line.split(maxsplit=1)[1] if len(line.split()) > 1 else ''
                                            break
                                    # Parse new value from new_value block
                                    new_val = ''
                                    for line in track['new_value'].split('\n'):
                                        if line.strip().startswith('WERT') or line.strip().startswith('TEXT'):
                                            new_val = line.split(maxsplit=1)[1] if len(line.split()) > 1 else ''
                                            break
                                    logs.append(('param_update', param_name, old_val, new_val, os.path.basename(fp)))
                else:
                    logs.append("No files required changes (values already correct).")

                self.root.after(0, lambda: self._finish_update(logs, matching_files, all_updated))
            except Exception as e:
                logs.append(f"Error during update: {e}")
                self.root.after(0, lambda: self._finish_update(logs, [], []))

        threading.Thread(target=worker, daemon=True).start()

    def start_remove(self):
        """Remove parameters from files."""
        if self._busy:
            messagebox.showinfo("Info", "Task đang chạy, vui lòng đợi.")
            return
        if not self.validate_inputs():
            return
        
        result = messagebox.askyesno(
            "Confirm Removal",
            f"⚠️ WARNING: This will PERMANENTLY DELETE all parameter blocks listed in the config file from {len(self.selected_directories)} directories.\n\n"
            f"This action cannot be undone!\n\n"
            f"Do you want to continue?",
            icon='warning'
        )
        if not result:
            return

        self.results_text.delete(1.0, tk.END)
        self.log_message("Starting parameter removal...")
        self._set_busy(True)

        def worker():
            logs = []
            try:
                ext = self.extension_var.get()
                config_path = self.list_file_combo.get().strip()
                config_parameters = read_config_file(config_path)
                if not config_parameters:
                    logs.append("No valid parameters found in config file.")
                    self.root.after(0, lambda: self._finish_remove(logs, [], []))
                    return

                logs.append(f"⚠️ Parameters to REMOVE: {len(config_parameters)}")
                for p in config_parameters.keys():
                    logs.append(f"  - {p}")
                
                # Find matching files
                logs.append(f"\nScanning {len(self.selected_directories)} directories...")
                matching_files = []
                for directory in self.selected_directories:
                    found_files = find_files_with_parameters(directory, config_parameters, ext)
                    matching_files.extend(found_files)
                
                if not matching_files:
                    logs.append("\nNo files found with the specified parameters.")
                    self.root.after(0, lambda: self._finish_remove(logs, [], []))
                    return

                logs.append(f"\nFound {len(matching_files)} files containing target parameters.")
                logs.append(f"\n🗑️ Removing parameters from files...")
                
                # Remove parameters
                removed_files, removal_tracking = remove_parameters_from_files(
                    matching_files, list(config_parameters.keys()), track_removals=True
                )
                
                # Store removal history
                self.removal_history = removal_tracking

                logs.append(f"\n✓ Successfully processed {len(matching_files)} files.")
                if removed_files:
                    logs.append(f"✓ Modified {len(removed_files)} files:")
                    grouped = {}
                    for fp in removed_files:
                        grouped.setdefault(os.path.dirname(fp), []).append(fp)
                    for d, file_paths in grouped.items():
                        logs.append(f"\nDirectory: {d}")
                        for fp in file_paths:
                            logs.append(f"  - {os.path.basename(fp)}")
                            # Show what was removed
                            for track in removal_tracking:
                                if track['file'] == fp:
                                    logs.append(('param_removed', track['parameter'], os.path.basename(fp)))
                else:
                    logs.append("No files were modified (parameters not found).")

                self.root.after(0, lambda: self._finish_remove(logs, matching_files, removed_files))
            except Exception as e:
                logs.append(f"Error during removal: {e}")
                import traceback
                logs.append(traceback.format_exc())
                self.root.after(0, lambda: self._finish_remove(logs, [], []))

        threading.Thread(target=worker, daemon=True).start()

    def _finish_remove(self, logs, all_matching, all_removed):
        for item in logs:
            if isinstance(item, tuple) and item[0] == 'param_removed':
                _, param_name, file_name = item
                timestamp = f"{datetime.now().strftime('%H:%M:%S')} - "
                self.results_text.insert(tk.END, timestamp)
                self.results_text.insert(tk.END, f"      🗑️ Removed: ", 'danger')
                self.results_text.insert(tk.END, f"{param_name}", 'param_name')
                self.results_text.insert(tk.END, f" from {file_name}\n", None)
                self.results_text.see(tk.END)
            else:
                self.log_message(item)
        
        if all_removed:
            removed_count = sum(len(self.removal_history) for _ in [None])
            # Clear cache to reflect updated files
            self.file_contents_cache = {}
            self.matching_files_cache = []
            messagebox.showinfo("Success", 
                f"Removed parameters from {len(all_removed)} files.\n"
                f"Total parameters removed: {len(self.removal_history)}")
        else:
            messagebox.showinfo("Info", "No parameters were removed.")
        self._set_busy(False)

    def _finish_update(self, logs, all_matching, all_updated):
        for item in logs:
            if isinstance(item, tuple) and item[0] == 'param_update':
                # Special handling for parameter update logs
                _, param_name, old_val, new_val, file_name = item
                timestamp = f"{datetime.now().strftime('%H:%M:%S')} - "
                self.results_text.insert(tk.END, timestamp)
                self.results_text.insert(tk.END, f"      ✓ Updated: ", 'success')
                self.results_text.insert(tk.END, f"{param_name}", 'param_name')
                self.results_text.insert(tk.END, f" [{old_val}]", None)
                self.results_text.insert(tk.END, f" \u2192 ", 'diff_highlight')
                self.results_text.insert(tk.END, f"[{new_val}]\\n", 'success')
                self.results_text.see(tk.END)
            else:
                self.log_message(item)
        if all_updated:
            # Clear cache to reflect updated files
            self.file_contents_cache = {}
            self.matching_files_cache = []
            messagebox.showinfo("Success", f"Updated {len(all_updated)} files.")
        else:
            messagebox.showinfo("Info", "No files updated.")
        self._set_busy(False)

    def _build_target_block(self, param_block, target_value):
        """Build target block by replacing WERT/TEXT value in parameter block"""
        if not param_block:
            return target_value
        
        lines = param_block.split('\n')
        new_lines = []
        for line in lines:
            stripped = line.strip()
            if stripped.startswith('WERT') or stripped.startswith('TEXT'):
                # Replace with new value, keeping indentation and keyword
                indentation = line[:len(line) - len(line.lstrip())]
                keyword = 'WERT' if stripped.startswith('WERT') else 'TEXT'
                new_lines.append(f"{indentation}{keyword} {target_value}")
            else:
                new_lines.append(line)
        return '\n'.join(new_lines)
    
    def generate_sheet_current_parameters(self, wb, file_params_cache, config_parameters):
        """Generate Sheet 1: Parameters to Update with Current and Target Values"""
        ws = wb.create_sheet("Parameters to Update", 0)
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        found_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        target_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Light yellow
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers - Show Original Value (before update) and Target Value
        headers = ["Parameter Name", "Original Value (Before Update)", "Target Value (Config)", "File Name", "Folder Path"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30  # Parameter Name
        ws.column_dimensions['B'].width = 60  # Original Value
        ws.column_dimensions['C'].width = 60  # Target Value 
        ws.column_dimensions['D'].width = 25  # File Name
        ws.column_dimensions['E'].width = 40  # Folder Path
        
        # Get all unique directories from selected_directories
        all_directories = set()
        for file_path in file_params_cache.keys():
            folder_path = os.path.dirname(file_path)
            all_directories.add(folder_path)
        
        # Also add directories that were selected but may have no matching files
        if hasattr(self, 'selected_directories'):
            for dir_path in self.selected_directories:
                all_directories.add(dir_path)
        
        # First pass: Build reference templates for each parameter from actual files
        param_templates = {}  # {param_name: sample_block}
        for file_path, all_params in file_params_cache.items():
            for param_detail in all_params:
                param_name = param_detail['name']
                if param_name in config_parameters and param_name not in param_templates:
                    # Store the first occurrence of each parameter as template
                    param_templates[param_name] = param_detail['block']
        
        # Populate data - ALL parameters from config for ALL directories
        row_num = 2
        not_found_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")  # Gray for missing
        
        for param_name, target_value in config_parameters.items():
            for folder_path in sorted(all_directories):
                # Find all files in this folder that are in cache
                files_in_folder = [fp for fp in file_params_cache.keys() if os.path.dirname(fp) == folder_path]
                
                if not files_in_folder:
                    # No files in cache for this folder - use template from other folders if available
                    if param_name in param_templates:
                        # Use actual parameter structure as template
                        target_block_structure = self._build_target_block(param_templates[param_name], target_value)
                    else:
                        # No template available anywhere - just show value
                        target_block_structure = target_value
                    
                    row_data = [param_name, "NA - No files checked in this folder", target_block_structure, "NA", folder_path]
                    ws.append(row_data)
                    
                    # Apply special formatting for folders with no files
                    for col_idx, cell in enumerate(ws[row_num], 1):
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.fill = not_found_fill
                        
                        if col_idx == 1:  # Parameter Name
                            cell.font = Font(bold=True, color="666666")
                        elif col_idx == 2:  # Original Value - "NA"
                            cell.font = Font(italic=True, color="666666")
                        elif col_idx == 3:  # Target Value - full structure
                            cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                            cell.font = Font(bold=True, color="D97706")
                        else:  # File Name, Folder Path
                            cell.font = Font(italic=True, color="666666")
                    
                    row_num += 1
                else:
                    # Check each file in this folder
                    param_found_in_folder = False
                    
                    for file_path in files_in_folder:
                        all_params = file_params_cache[file_path]
                        file_name = os.path.basename(file_path)
                        
                        # Look for the parameter in this file
                        param_found = False
                        param_block = None
                        
                        for param_detail in all_params:
                            if param_detail['name'] == param_name:
                                param_found = True
                                param_found_in_folder = True
                                param_block = param_detail['block']
                                break
                        
                        if param_found:
                            # Build target block using helper method
                            target_block = self._build_target_block(param_block, target_value)
                            
                            # Extract actual values for comparison
                            original_value = ""
                            target_value_clean = target_value
                            
                            for line in param_block.split('\n'):
                                stripped = line.strip()
                                if stripped.startswith('WERT') or stripped.startswith('TEXT'):
                                    parts = stripped.split(maxsplit=1)
                                    if len(parts) > 1:
                                        original_value = parts[1]
                                    break
                            
                            # Check if values differ
                            values_differ = (original_value.strip() != target_value_clean.strip())
                            
                            row_data = [param_name, param_block, target_block, file_name, folder_path]
                            ws.append(row_data)
                            
                            # Apply formatting with conditional highlighting
                            for col_idx, cell in enumerate(ws[row_num], 1):
                                cell.border = border
                                cell.alignment = Alignment(vertical="top", wrap_text=True)
                                
                                if values_differ:
                                    # Highlight entire row if values differ
                                    if col_idx == 1:  # Parameter Name - highlight in red
                                        cell.font = Font(bold=True, color="C00000")
                                    elif col_idx == 2:  # Original Value - red background
                                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                        cell.font = Font(bold=True, color="9C0006")
                                    elif col_idx == 3:  # Target Value - yellow/orange highlight
                                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                                        cell.font = Font(bold=True, color="9C5700")
                                else:
                                    # Values match - use green
                                    if col_idx == 2 or col_idx == 3:
                                        cell.fill = found_fill  # Light green
                            
                            row_num += 1
                    
                    # If parameter not found in any file in this folder
                    if not param_found_in_folder:
                        # Use template from other folders if available
                        if param_name in param_templates:
                            # Use actual parameter structure as template
                            target_block_structure = self._build_target_block(param_templates[param_name], target_value)
                        else:
                            # No template available anywhere - just show value
                            target_block_structure = target_value
                        
                        row_data = [param_name, "NA - Parameter not found in folder", target_block_structure, "NA", folder_path]
                        ws.append(row_data)
                        
                        # Apply special formatting for missing parameters
                        for col_idx, cell in enumerate(ws[row_num], 1):
                            cell.border = border
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
                            
                            if col_idx == 1:  # Parameter Name - Gray bold
                                cell.font = Font(bold=True, color="666666")
                                cell.fill = not_found_fill
                            elif col_idx == 2:  # Original Value - Gray with "NA"
                                cell.fill = not_found_fill
                                cell.font = Font(italic=True, color="666666")
                            elif col_idx == 3:  # Target Value - Light orange to show what would be added
                                cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
                                cell.font = Font(bold=True, color="D97706")
                            else:
                                cell.fill = not_found_fill
                                cell.font = Font(italic=True, color="666666")
                        
                        row_num += 1
        
        ws.freeze_panes = "A2"
    
    def generate_sheet_folder_structure(self, wb, file_params_cache, config_parameters):
        """Generate Sheet 2: Folder and File Structure"""
        ws = wb.create_sheet("Folder Structure", 1)
        
        # Define styles
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        folder_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers
        headers = ["Folder Path", "File Name", "Full File Path", "Parameters Found", "Parameter Count"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        
        # Group files by folder
        folder_dict = {}
        for file_path in file_params_cache.keys():
            folder_path = os.path.dirname(file_path)
            if folder_path not in folder_dict:
                folder_dict[folder_path] = []
            folder_dict[folder_path].append(file_path)
        
        # Populate data
        row_num = 2
        for folder_path in sorted(folder_dict.keys()):
            for file_path in sorted(folder_dict[folder_path]):
                file_name = os.path.basename(file_path)
                all_params = file_params_cache[file_path]
                param_names = [p['name'] for p in all_params]
                param_count = len(param_names)
                params_str = ", ".join(param_names[:5])
                if param_count > 5:
                    params_str += f"... (+{param_count - 5} more)"
                
                row_data = [folder_path, file_name, file_path, params_str, param_count]
                ws.append(row_data)
                
                for cell in ws[row_num]:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if cell.column == 1:
                        cell.fill = folder_fill
                        cell.font = Font(bold=True)
                
                row_num += 1
        
        ws.freeze_panes = "A2"
    
    def generate_sheet_update_comparison(self, wb, config_parameters):
        """Generate Sheet 3: Update Comparison with Config Values"""
        ws = wb.create_sheet("Update Comparison", 2)
        
        # Define styles
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        updated_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Blue
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        #Set headers
        headers = ["Folder", "File", "Parameter", "Old Value", "New Value", "Status"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 40  # Folder
        ws.column_dimensions['B'].width = 25  # File
        ws.column_dimensions['C'].width = 30  # Parameter
        ws.column_dimensions['D'].width = 60  # Old Value
        ws.column_dimensions['E'].width = 60  # New Value
        ws.column_dimensions['F'].width = 12  # Status
        
        # Populate data from update history
        row_num = 2
        for entry in self.update_history:
            folder = entry.get('folder', '')
            file_name = os.path.basename(entry.get('file', ''))
            parameter = entry.get('parameter', '')
            old_value = entry.get('old_value', '')
            new_value = entry.get('new_value', '')
            status = entry.get('status', '')
            
            row_data = [folder, file_name, parameter, old_value, new_value, status]
            ws.append(row_data)
            
            # Apply formatting
            for cell in ws[row_num]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = updated_fill
            
            row_num += 1
        
        # If no update history, show message
        if not self.update_history:
            ws.append(["No update operations recorded yet.", "", "", "", "", ""])
            ws['A2'].font = Font(italic=True, color="666666")
        
        ws.freeze_panes = "A2"
    
    def generate_sheet_removal_tracking(self, wb):
        """Generate Sheet 4: Parameter Removal Tracking"""
        ws = wb.create_sheet("Parameter Removals", 3)
        
        # Define styles
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Light red
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers
        headers = ["Folder", "File", "Parameter", "Removed Block", "Status"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 40  # Folder
        ws.column_dimensions['B'].width = 25  # File
        ws.column_dimensions['C'].width = 30  # Parameter
        ws.column_dimensions['D'].width = 80  # Removed Block
        ws.column_dimensions['E'].width = 12  # Status
        
        # Populate data from removal history
        row_num = 2
        for entry in self.removal_history:
            folder = entry.get('folder', '')
            file_name = os.path.basename(entry.get('file', ''))
            parameter = entry.get('parameter', '')
            removed_block = entry.get('removed_block', '')
            status = entry.get('status', '')
            
            row_data = [folder, file_name, parameter, removed_block, status]
            ws.append(row_data)
            
            # Apply formatting
            for cell in ws[row_num]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.fill = removed_fill
            
            row_num += 1
        
        # If no removal history, show message
        if not self.removal_history:
            ws.append(["No removal operations recorded yet.", "", "", "", ""])
            ws['A2'].font = Font(italic=True, color="666666")
        
        ws.freeze_panes = "A2"
    
    def export_excel_report(self):
        """Export comprehensive Excel report with 3 sheets"""
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", 
                "openpyxl library is not installed.\\n\\n"
                "Please install it using:\\npip install openpyxl")
            return
        
        # Check if config file is selected
        config_path = self.list_file_combo.get().strip()
        if not config_path or not os.path.isfile(config_path):
            messagebox.showerror("Error", "Please select a valid config file first.")
            return
        
        # Check if directories are selected
        if not self.selected_directories:
            messagebox.showerror("Error", "Please select directories first.")
            return
        
        try:
            self.log_message("Starting Excel report generation...")
            self.root.update()
            
            # Load config parameters
            self.config_parameters = read_config_file(config_path)
            if not self.config_parameters:
                messagebox.showerror("Error", "No valid parameters found in config file.")
                return
            
            self.log_message(f"Loaded {len(self.config_parameters)} parameters from config")
            
            # Use cached original values if available (from before update), otherwise read files now
            if self.file_contents_cache:
                self.log_message(f"Using cached original values from {len(self.file_contents_cache)} files")
                file_params_cache = self.file_contents_cache
                file_count = len(file_params_cache)
            else:
                # Find only files that contain the parameters to update
                self.log_message(f"Finding files with target parameters in {len(self.selected_directories)} directories...")
                self.root.update()
                
                ext = self.extension_var.get()
                matching_files = []
                
                for directory in self.selected_directories:
                    files_in_dir = find_files_with_parameters(directory, self.config_parameters, ext)
                    matching_files.extend(files_in_dir)
                
                self.log_message(f"Found {len(matching_files)} files containing target parameters")
                self.root.update()
                
                # Pre-extract parameters only from matching files
                file_params_cache = {}
                file_count = 0
                
                for file_path in matching_files:
                    file_params_cache[file_path] = extract_all_parameters_from_file(file_path)
                    file_count += 1
                    
                    if file_count % 10 == 0:
                        self.log_message(f"  Processed {file_count} files...")
                        self.root.update()
            
            total_params = sum(len(params) for params in file_params_cache.values())
            self.log_message(f"Extracted {total_params} total parameters from {file_count} files")
            self.root.update()
            
            # Create Report folder
            report_folder = os.path.join(os.path.dirname(__file__), 'Report')
            if not os.path.exists(report_folder):
                os.makedirs(report_folder)
                self.log_message(f"Created Report folder: {report_folder}")
            
            # Create workbook
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Generate all sheets
            self.log_message("Generating Sheet 1: Current Parameters...")
            self.root.update()
            self.generate_sheet_current_parameters(wb, file_params_cache, self.config_parameters)
            
            self.log_message("Generating Sheet 2: Folder Structure...")
            self.root.update()
            self.generate_sheet_folder_structure(wb, file_params_cache, self.config_parameters)
            
            self.log_message("Generating Sheet 3: Update Comparison...")
            self.root.update()
            self.generate_sheet_update_comparison(wb, self.config_parameters)
            
            self.log_message("Generating Sheet 4: Parameter Removals...")
            self.root.update()
            self.generate_sheet_removal_tracking(wb)
            
            # Save workbook
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"Update_Report_{timestamp}.xlsx"
            excel_path = os.path.join(report_folder, excel_filename)
            
            self.log_message("Saving Excel file...")
            self.root.update()
            wb.save(excel_path)
            
            self.log_message(f"Excel report saved successfully!")
            self.log_message(f"Location: {excel_path}")
            
            messagebox.showinfo("Success", 
                f"Excel report generated successfully!\\n\\n"
                f"File: {excel_filename}\\n"
                f"Location: {report_folder}\\n\\n"
                f"Total files analyzed: {file_count}\\n"
                f"Parameters tracked: {len(self.config_parameters)}\\n"
                f"Update operations: {len(self.update_history)}\\n"
                f"Removal operations: {len(self.removal_history)}")
            
            # Optional: Open folder
            if messagebox.askyesno("Open Folder", "Do you want to open the Report folder?"):
                os.startfile(report_folder)
                
        except Exception as e:
            error_msg = f"Error generating Excel report: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("Export Error", error_msg)

if __name__ == "__main__":
    root = tk.Tk()
    app = ParameterUpdateTool(root)
    root.mainloop()