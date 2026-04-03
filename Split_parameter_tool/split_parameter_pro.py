import re
import os
import json
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from tkinter.scrolledtext import ScrolledText
import threading
from datetime import datetime
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not installed. Excel export will not be available.")

def read_file(file_path):
    with open(file_path, 'r') as file:
        return file.read()

def write_file(file_path, content):
    with open(file_path, 'w') as file:
        file.write(content)

def read_parameter_list(list_file_path):
    with open(list_file_path, 'r') as file:
        return [line.strip() for line in file.readlines() if line.strip()]

def find_specific_kennlinie(content, target_name):
    patterns = [
        r'(KENNLINIE\s+' + re.escape(target_name) + r'\s+[\s\S]+?END)',
        r'(KENNFELD\s+' + re.escape(target_name) + r'\s+[\s\S]+?END)',
        r'(FESTWERT\s+' + re.escape(target_name) + r'\s+[\s\S]+?END)',
        r'(GRUPPENKENNLINIE\s+' + re.escape(target_name) + r'\s+[\s\S]+?END)',
    ]
    
    for pattern in patterns:
        matches = re.findall(pattern, content)
        if matches:
            return matches[0]
    
    generic_pattern = r'([A-Z]+\s+' + re.escape(target_name) + r'\s+[\s\S]+?END)'
    matches = re.findall(generic_pattern, content)
    if matches:
        return matches[0]
    
    return None

def clone_parameter(content, target_name, suffix, position="Suffix"):
    block = find_specific_kennlinie(content, target_name)
    
    if not block:
        return None

    if position == "Prefix":
        # If suffix ends with _ or target_name starts with _, don't add extra _
        if suffix.endswith('_') or target_name.startswith('_'):
            new_target_name = f"{suffix}{target_name}"
        else:
            new_target_name = f"{suffix}_{target_name}"
    elif position == "Replace":
        new_target_name = suffix  # Replace entire parameter name with suffix value
    else:  # Suffix
        # If target_name ends with _ or suffix starts with _, don't add extra _
        if target_name.endswith('_') or suffix.startswith('_'):
            new_target_name = f"{target_name}{suffix}"
        else:
            new_target_name = f"{target_name}_{suffix}"
    
    lines = block.split('\n')
    first_line_parts = lines[0].split(maxsplit=2)
    
    if len(first_line_parts) < 2:
        return None
    
    param_type = first_line_parts[0]
    new_first_line = f"{param_type} {new_target_name}"
    
    if len(first_line_parts) > 2:
        new_first_line += f" {first_line_parts[2]}"
        
    lines[0] = new_first_line
    
    cleaned_lines = [lines[0]]
    for line in lines[1:]:
        if line.strip():
            cleaned_lines.append(line)
    
    new_block = '\n'.join(cleaned_lines)
    return new_block.strip()

def extract_parameter_type(block):
    """Extract parameter type (KENNLINIE, KENNFELD, FESTWERT, etc.) from parameter block."""
    if not block:
        return "Unknown"
    
    first_line = block.split('\n')[0].strip()
    parts = first_line.split(maxsplit=1)
    
    if len(parts) > 0:
        return parts[0]
    return "Unknown"

def extract_parameter_value(block):
    """Extract WERT or TEXT value from parameter block."""
    if not block:
        return ""
    
    lines = block.split('\n')
    for line in lines:
        stripped = line.strip()
        if stripped.startswith('WERT') or stripped.startswith('TEXT'):
            parts = stripped.split(maxsplit=1)
            if len(parts) > 1:
                return parts[1]
    return ""

def parse_parameter_details(content, param_name):
    """Get full parameter details including type and value."""
    block = find_specific_kennlinie(content, param_name)
    
    if not block:
        return None
    
    param_type = extract_parameter_type(block)
    param_value = extract_parameter_value(block)
    
    return {
        'name': param_name,
        'type': param_type,
        'value': param_value,
        'block': block
    }

def extract_all_parameters(content):
    """Extract all parameters from DCM content efficiently in one pass."""
    parameters = []
    keywords = ("KENNLINIE", "KENNFELD", "FESTWERT", "GRUPPENKENNLINIE")
    
    lines = content.split('\n')
    i = 0
    
    while i < len(lines):
        stripped = lines[i].strip()
        
        # Check if this line starts a parameter definition
        if stripped.startswith(keywords):
            parts = stripped.split()
            if len(parts) > 1:
                param_name = parts[1]
                param_type = parts[0]
                
                # Collect the entire block until END
                block_lines = [lines[i]]
                i += 1
                
                while i < len(lines):
                    block_lines.append(lines[i])
                    if lines[i].strip() == 'END':
                        break
                    i += 1
                
                block = '\n'.join(block_lines)
                
                # Extract value from block
                param_value = ""
                for line in block_lines:
                    line_stripped = line.strip()
                    if line_stripped.startswith('WERT') or line_stripped.startswith('TEXT'):
                        value_parts = line_stripped.split(maxsplit=1)
                        if len(value_parts) > 1:
                            param_value = value_parts[1]
                        break
                
                # Add to parameters list
                parameters.append({
                    'name': param_name,
                    'type': param_type,
                    'value': param_value,
                    'block': block
                })
        
        i += 1
    
    return parameters

class MultiDirectoryDialog:
    def __init__(self, parent):
        self.parent = parent
        self.selected_directories = []
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Select Multiple Directories")
        self.dialog.geometry("700x700")
        self.dialog.grab_set()
        
        # Load parent folder history
        self.config_file = os.path.join(os.path.dirname(__file__), 'split_parameter_config.json')
        self.parent_folder_history = []
        self.load_parent_history()
        
        self.setup_dialog()
        
    def setup_dialog(self):
        main_frame = tk.Frame(self.dialog, padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        instruction_text = "Select multiple directories:\n• Use 'Browse Parent Folder' to navigate to a parent directory\n• Use Ctrl+A to select all folders\n• Use Ctrl+Click or Shift+Click for multiple selection"
        tk.Label(main_frame, text=instruction_text, font=('Arial', 10), 
                justify=tk.LEFT, fg="blue").pack(anchor='w', pady=(0, 15))
        
        path_frame = tk.Frame(main_frame)
        path_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(path_frame, text="Parent Folder:", font=('Arial', 9, 'bold')).pack(side=tk.LEFT)
        
        # Combobox for parent folder with history
        path_combo_frame = tk.Frame(path_frame)
        path_combo_frame.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 10))
        
        self.current_path_var = tk.StringVar(value="")
        self.path_combo = ttk.Combobox(path_combo_frame, textvariable=self.current_path_var, 
                                       font=('Arial', 9))
        self.path_combo['values'] = self.parent_folder_history
        if self.parent_folder_history:
            self.current_path_var.set(self.parent_folder_history[0])
        self.path_combo.pack(fill=tk.X)
        self.path_combo.bind('<<ComboboxSelected>>', self.on_path_selected)
        
        # Browse button
        tk.Button(path_frame, text="Browse", command=self.browse_parent_folder, 
                 width=10, bg='lightgreen').pack(side=tk.LEFT)
        
        tk.Label(main_frame, text="Available Folders in Parent Directory:", 
                font=('Arial', 10, 'bold')).pack(anchor='w', pady=(10, 5))
        
        available_frame = tk.Frame(main_frame)
        available_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.available_listbox = tk.Listbox(available_frame, selectmode=tk.EXTENDED, height=8)
        available_scrollbar = tk.Scrollbar(available_frame, orient=tk.VERTICAL, 
                                         command=self.available_listbox.yview)
        self.available_listbox.configure(yscrollcommand=available_scrollbar.set)
        
        self.available_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        available_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.available_listbox.bind('<Control-a>', self.select_all_available)
        self.available_listbox.bind('<Return>', self.add_selected_directories)
        self.available_listbox.bind('<Double-Button-1>', self.add_selected_directories)
        
        # Load initial directory if history exists
        if self.parent_folder_history:
            self.load_subdirectories(self.parent_folder_history[0])
        
        # TRANSFER BUTTONS - These were missing!
        transfer_frame = tk.Frame(main_frame)
        transfer_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Button(transfer_frame, text="Add Selected Folders →", command=self.add_selected_directories, 
                 width=20, bg='lightblue').pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(transfer_frame, text="Add All Folders →", command=self.add_all_directories, 
                 width=15).pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Label(main_frame, text="Selected Directories:", font=('Arial', 10, 'bold')).pack(anchor='w', pady=(10, 5))
        
        selected_frame = tk.Frame(main_frame)
        selected_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))
        
        self.selected_listbox = tk.Listbox(selected_frame, selectmode=tk.EXTENDED, height=6)
        selected_scrollbar = tk.Scrollbar(selected_frame, orient=tk.VERTICAL, 
                                        command=self.selected_listbox.yview)
        self.selected_listbox.configure(yscrollcommand=selected_scrollbar.set)
        
        self.selected_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        selected_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.selected_listbox.bind('<Control-a>', self.select_all_selected)
        self.selected_listbox.bind('<Delete>', self.remove_selected_directories)
        
        mgmt_frame = tk.Frame(main_frame)
        mgmt_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Button(mgmt_frame, text="Remove Selected", command=self.remove_selected_directories, 
                 width=15).pack(side=tk.LEFT, padx=(0, 5))
        tk.Button(mgmt_frame, text="Clear All", command=self.clear_all, width=10).pack(side=tk.LEFT)
        
        self.count_label = tk.Label(mgmt_frame, text="Selected: 0 directories", fg="blue", 
                                   font=('Arial', 10, 'bold'))
        self.count_label.pack(side=tk.RIGHT)
        
        dialog_button_frame = tk.Frame(main_frame)
        dialog_button_frame.pack(fill=tk.X, pady=(10, 0))
        
        tk.Button(dialog_button_frame, text="OK", command=self.ok_clicked, 
                 width=10, bg='lightgreen', font=('Arial', 10, 'bold')).pack(side=tk.RIGHT, padx=(5, 0))
        tk.Button(dialog_button_frame, text="Cancel", command=self.cancel_clicked, 
                 width=10, font=('Arial', 10)).pack(side=tk.RIGHT)
        
        self.center_dialog()
        
    def center_dialog(self):
        self.dialog.transient(self.parent)
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (self.dialog.winfo_width() // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (self.dialog.winfo_height() // 2)
        self.dialog.geometry(f"+{x}+{y}")
    
    def load_parent_history(self):
        """Load parent folder history from config file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.parent_folder_history = config.get('parent_folder_history', [])
        except Exception as e:
            print(f"Error loading parent folder history: {e}")
            self.parent_folder_history = []
    
    def save_parent_history(self):
        """Save parent folder history to config file"""
        try:
            config = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            config['parent_folder_history'] = self.parent_folder_history
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving parent folder history: {e}")
    
    def on_path_selected(self, event=None):
        """Handle selection from dropdown"""
        selected_path = self.current_path_var.get()
        if selected_path and os.path.exists(selected_path):
            self.load_subdirectories(selected_path)
        
    def browse_parent_folder(self):
        parent_dir = filedialog.askdirectory(title="Select Parent Directory")
        if parent_dir:
            self.current_path_var.set(parent_dir)
            self.load_subdirectories(parent_dir)
            
            # Save to history
            if parent_dir not in self.parent_folder_history:
                self.parent_folder_history.insert(0, parent_dir)
                self.parent_folder_history = self.parent_folder_history[:10]  # Keep last 10
                self.path_combo['values'] = self.parent_folder_history
                self.save_parent_history()
            
    def load_subdirectories(self, parent_dir):
        self.available_listbox.delete(0, tk.END)
        try:
            items = os.listdir(parent_dir)
            directories = []
            for item in items:
                item_path = os.path.join(parent_dir, item)
                if os.path.isdir(item_path):
                    directories.append(item_path)
            
            directories.sort()
            for directory in directories:
                self.available_listbox.insert(tk.END, directory)
                
        except Exception as e:
            messagebox.showerror("Error", f"Error reading directory: {str(e)}")
    
    def select_all_available(self, event=None):
        self.available_listbox.select_set(0, tk.END)
        return 'break'
    
    def select_all_selected(self, event=None):
        self.selected_listbox.select_set(0, tk.END)
        return 'break'
        
    def add_selected_directories(self, event=None):
        selections = self.available_listbox.curselection()
        for index in selections:
            directory = self.available_listbox.get(index)
            if directory not in self.selected_directories:
                self.selected_directories.append(directory)
                self.selected_listbox.insert(tk.END, directory)
        self.update_count()
        
    def add_all_directories(self):
        for i in range(self.available_listbox.size()):
            directory = self.available_listbox.get(i)
            if directory not in self.selected_directories:
                self.selected_directories.append(directory)
                self.selected_listbox.insert(tk.END, directory)
        self.update_count()
        
    def remove_selected_directories(self, event=None):
        selections = self.selected_listbox.curselection()
        for index in reversed(selections):
            directory = self.selected_listbox.get(index)
            self.selected_listbox.delete(index)
            self.selected_directories.remove(directory)
        self.update_count()
        
    def clear_all(self):
        self.selected_listbox.delete(0, tk.END)
        self.selected_directories.clear()
        self.update_count()
        
    def update_count(self):
        count = len(self.selected_directories)
        self.count_label.config(text=f"Selected: {count} directories")
        
    def ok_clicked(self):
        self.dialog.destroy()
        
    def cancel_clicked(self):
        self.selected_directories.clear()
        self.dialog.destroy()

class ParameterClonerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Multi-File Parameter Cloner Tool")
        self.root.geometry("900x1000")
        self.root.minsize(800, 700)  # Minimum window size
        
        # Modern color scheme
        self.colors = {
            'primary': '#2196F3',
            'secondary': '#1976D2',
            'success': '#4CAF50',
            'warning': '#FF9800',
            'danger': '#F44336',
            'bg': '#f0f0f0',
            'panel': '#ffffff',
            'text': '#212121',
            'text_secondary': '#757575'
        }
        
        self.parameter_files = []
        self.file_contents = {}
        self.list_file_path = ""
        self.parameter_list = []
        self.suffix_array = ["Super"]
        self.parameter_widgets = {}
        
        # Config file and history
        self.config_file = os.path.join(os.path.dirname(__file__), 'split_parameter_config.json')
        self.list_file_history = []
        self.scan_directory_history = []
        self.load_config()
        
        # Excel report tracking
        self.cloning_history = []  # Track all cloning operations for Excel report
        
        self.setup_gui()
    
    def setup_gui(self):
        # Header frame
        header_frame = tk.Frame(self.root, bg=self.colors['primary'], height=60)
        header_frame.pack(fill=tk.X, side=tk.TOP)
        header_frame.pack_propagate(False)
        
        title_label = tk.Label(header_frame, text="⚙ Parameter Cloner Pro", 
                              font=('Segoe UI', 16, 'bold'), 
                              bg=self.colors['primary'], fg='white')
        title_label.pack(side=tk.LEFT, padx=20, pady=15)
        
        version_label = tk.Label(header_frame, text="v2.0", 
                               font=('Segoe UI', 9), 
                               bg=self.colors['primary'], fg='white')
        version_label.pack(side=tk.RIGHT, padx=20, pady=15)
        
        # Main content frame
        main_frame = tk.Frame(self.root, padx=25, pady=20, bg=self.colors['bg'])
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Configure grid weights for responsive resizing
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)  # Results section gets extra space
        
        # Parameter List File Section
        list_section = tk.LabelFrame(main_frame, text="📄 Parameter List File", 
                                    font=('Segoe UI', 10, 'bold'), 
                                    bg=self.colors['panel'], 
                                    fg=self.colors['text'],
                                    relief=tk.FLAT, 
                                    borderwidth=2)
        list_section.grid(row=0, column=0, columnspan=3, sticky='nsew', pady=(0, 15), padx=2)
        main_frame.columnconfigure(0, weight=1)
        
        list_file_frame = tk.Frame(list_section, bg=self.colors['panel'])
        list_file_frame.pack(fill=tk.X, padx=15, pady=15)
        list_file_frame.columnconfigure(0, weight=1)
        
        self.list_file_var = tk.StringVar(value="")
        self.list_file_combo = ttk.Combobox(list_file_frame, textvariable=self.list_file_var,
                                            font=('Segoe UI', 9))
        self.list_file_combo['values'] = self.list_file_history
        if self.list_file_history:
            self.list_file_var.set(self.list_file_history[0])
        self.list_file_combo.grid(row=0, column=0, sticky='ew', padx=(0, 10), ipady=6)
        
        browse_btn = tk.Button(list_file_frame, text="📁 Browse", 
                              command=self.select_list_file,
                              font=('Segoe UI', 8),
                              bg=self.colors['primary'], 
                              fg='white',
                              relief=tk.FLAT,
                              cursor='hand2',
                              padx=12, pady=5)
        browse_btn.grid(row=0, column=1)
        browse_btn.bind('<Enter>', lambda e: browse_btn.config(bg=self.colors['secondary']))
        browse_btn.bind('<Leave>', lambda e: browse_btn.config(bg=self.colors['primary']))
        
        # Parameter Files Section
        files_section = tk.LabelFrame(main_frame, text="📂 Parameter Files Selection", 
                                     font=('Segoe UI', 10, 'bold'), 
                                     bg=self.colors['panel'], 
                                     fg=self.colors['text'],
                                     relief=tk.FLAT, 
                                     borderwidth=2)
        files_section.grid(row=1, column=0, columnspan=3, sticky='nsew', pady=(0, 15), padx=2)
        main_frame.rowconfigure(1, weight=1)  # Give some weight to files section
        
        files_frame = tk.Frame(files_section, bg=self.colors['panel'])
        files_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        files_frame.columnconfigure(0, weight=1)
        
        listbox_frame = tk.Frame(files_frame, bg=self.colors['panel'])
        listbox_frame.grid(row=0, column=0, sticky='nsew', padx=(0, 15))
        listbox_frame.columnconfigure(0, weight=1)
        listbox_frame.rowconfigure(0, weight=1)
        
        self.file_listbox = tk.Listbox(listbox_frame, height=5, 
                                       font=('Segoe UI', 9),
                                       relief=tk.SOLID,
                                       borderwidth=1,
                                       selectbackground=self.colors['primary'],
                                       selectforeground='white')
        self.file_listbox.grid(row=0, column=0, sticky='nsew')
        
        file_scrollbar = tk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        file_scrollbar.grid(row=0, column=1, sticky='ns')
        self.file_listbox.configure(yscrollcommand=file_scrollbar.set)
        
        file_buttons_frame = tk.Frame(files_frame, bg=self.colors['panel'])
        file_buttons_frame.grid(row=0, column=1, sticky='n')
        
        btn_style = {'font': ('Segoe UI', 8), 'relief': tk.FLAT, 'cursor': 'hand2', 'width': 16, 'pady': 4}
        
        add_files_btn = tk.Button(file_buttons_frame, text="➕ Add Files", command=self.add_parameter_files,
                                 bg=self.colors['primary'], fg='white', **btn_style)
        add_files_btn.pack(pady=(0, 6))
        
        scan_single_btn = tk.Button(file_buttons_frame, text="🔍 Scan Directory", 
                                   command=self.scan_directory_for_parameters,
                                   bg=self.colors['secondary'], fg='white', **btn_style)
        scan_single_btn.pack(pady=(0, 6))
        
        scan_multi_btn = tk.Button(file_buttons_frame, text="🔍 Scan Multiple", 
                                  command=self.scan_multiple_directories,
                                  bg=self.colors['secondary'], fg='white', **btn_style)
        scan_multi_btn.pack(pady=(0, 6))
        
        remove_btn = tk.Button(file_buttons_frame, text="🗑️ Remove", 
                              command=self.remove_selected_file,
                              bg=self.colors['warning'], fg='white', **btn_style)
        remove_btn.pack(pady=(0, 6))
        
        clear_btn = tk.Button(file_buttons_frame, text="✖ Clear All", 
                             command=self.clear_all_files,
                             bg=self.colors['danger'], fg='white', **btn_style)
        clear_btn.pack()
        
        self.file_count_label = tk.Label(files_section, text="📊 Selected files: 0", 
                                         fg=self.colors['primary'], 
                                         bg=self.colors['panel'],
                                         font=('Segoe UI', 9, 'bold'))
        self.file_count_label.pack(side=tk.LEFT, padx=15, pady=(0, 10))
        
        # Suffix Configuration Section
        suffix_frame = tk.LabelFrame(main_frame, text="⚙ Suffix Configuration", 
                                    font=('Segoe UI', 10, 'bold'),
                                    bg=self.colors['panel'], 
                                    fg=self.colors['text'],
                                    relief=tk.FLAT, 
                                    borderwidth=2)
        suffix_frame.grid(row=2, column=0, columnspan=3, sticky='nsew', pady=(0, 15), padx=2)
        suffix_frame.columnconfigure(1, weight=1)
        
        tk.Label(suffix_frame, text="Define suffix array (up to 6):", 
                font=('Segoe UI', 9, 'bold'),
                bg=self.colors['panel'],
                fg=self.colors['text']).grid(row=0, column=0, sticky='w', padx=15, pady=15)
        
        suffix_entries_frame = tk.Frame(suffix_frame, bg=self.colors['panel'])
        suffix_entries_frame.grid(row=0, column=1, sticky='ew', padx=15, pady=15)
        
        self.suffix_entries = []
        for i in range(6):
            entry = tk.Entry(suffix_entries_frame, width=12,
                           font=('Segoe UI', 9),
                           relief=tk.SOLID,
                           borderwidth=1,
                           justify='center')
            entry.grid(row=0, column=i, padx=3, ipady=4)
            
            if i < len(self.suffix_array):
                entry.insert(0, self.suffix_array[i])
                
            self.suffix_entries.append(entry)
        
        # Cloning Options Section
        cloning_frame = tk.LabelFrame(main_frame, text="🔧 Cloning Options", 
                                     font=('Segoe UI', 10, 'bold'),
                                     bg=self.colors['panel'], 
                                     fg=self.colors['text'],
                                     relief=tk.FLAT, 
                                     borderwidth=2)
        cloning_frame.grid(row=3, column=0, columnspan=3, sticky='nsew', pady=(0, 15), padx=2)
        
        self.clone_option_var = tk.StringVar(value="add_suffix")
        
        options_inner_frame = tk.Frame(cloning_frame)
        options_inner_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Radiobutton(
            options_inner_frame, 
            text="Add new suffix", 
            variable=self.clone_option_var, 
            value="add_suffix",
            command=self.toggle_clone_options,
            font=('Arial', 9)
        ).grid(row=0, column=0, sticky='w', padx=5)
        
        tk.Radiobutton(
            options_inner_frame, 
            text="Replace suffix and split new parameter", 
            variable=self.clone_option_var, 
            value="replace_text",
            command=self.toggle_clone_options,
            font=('Arial', 9)
        ).grid(row=0, column=1, sticky='w', padx=5)
        
        self.add_suffix_frame = tk.Frame(cloning_frame)
        self.replace_text_frame = tk.Frame(cloning_frame)
        
        position_frame = tk.Frame(self.add_suffix_frame)
        position_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(position_frame, text="Position:", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w')
        
        self.position_var = tk.StringVar(value="Suffix")
        self.position_var.trace('w', self.on_position_change)
        
        tk.Radiobutton(position_frame, text="Prefix (add before name)", variable=self.position_var, value="Prefix", font=('Arial', 9)).grid(row=0, column=1, padx=5)
        tk.Radiobutton(position_frame, text="Suffix (add after name)", variable=self.position_var, value="Suffix", font=('Arial', 9)).grid(row=0, column=2, padx=5)
        tk.Radiobutton(position_frame, text="Replace (replace entire name)", variable=self.position_var, value="Replace", font=('Arial', 9)).grid(row=0, column=3, padx=5)
        
        # Replace mode frame (only shown when Replace position is selected)
        self.replace_mode_frame_suffix = tk.Frame(self.add_suffix_frame)
        
        tk.Label(self.replace_mode_frame_suffix, text="Replace mode:", font=('Arial', 9, 'bold')).grid(row=0, column=0, sticky='w', padx=5)
        
        self.replace_mode_suffix_var = tk.StringVar(value="clone_replace")
        
        tk.Radiobutton(
            self.replace_mode_frame_suffix,
            text="Replace directly (modify existing parameter)",
            variable=self.replace_mode_suffix_var,
            value="direct_replace",
            font=('Arial', 9)
        ).grid(row=0, column=1, sticky='w', padx=5)
        
        tk.Radiobutton(
            self.replace_mode_frame_suffix,
            text="Clone and replace (create new parameter)",
            variable=self.replace_mode_suffix_var,
            value="clone_replace",
            font=('Arial', 9)
        ).grid(row=0, column=2, sticky='w', padx=5)
        
        replace_inner_frame = tk.Frame(self.replace_text_frame)
        replace_inner_frame.pack(fill='x', padx=5, pady=5)
        
        tk.Label(replace_inner_frame, text="Find suffix:", font=('Arial', 9)).grid(row=0, column=0, sticky='w', padx=5)
        self.find_text_entry = tk.Entry(replace_inner_frame, width=20)
        self.find_text_entry.grid(row=0, column=1, padx=5)
        self.find_text_entry.insert(0, "Sport_1")
        
        tk.Label(replace_inner_frame, text="New suffix:", font=('Arial', 9)).grid(row=0, column=2, sticky='w', padx=5)
        self.replace_with_entry = tk.Entry(replace_inner_frame, width=20)
        self.replace_with_entry.grid(row=0, column=3, padx=5)
        self.replace_with_entry.insert(0, "Sport_3")
        
        self.toggle_clone_options()
        
        # Action Buttons
        button_frame = tk.Frame(main_frame, bg=self.colors['bg'])
        button_frame.grid(row=4, column=0, columnspan=3, pady=(10, 0))
        
        action_btn_style = {'font': ('Segoe UI', 9, 'bold'), 'relief': tk.FLAT, 'cursor': 'hand2', 'pady': 8, 'padx': 15}
        
        load_btn = tk.Button(button_frame, text="1. Load Files", command=self.load_files, 
                            bg=self.colors['primary'], fg='white', width=15, **action_btn_style)
        load_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        update_btn = tk.Button(button_frame, text="2. Update Array", command=self.update_suffix_array, 
                              bg=self.colors['warning'], fg='white', width=15, **action_btn_style)
        update_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        clone_btn = tk.Button(button_frame, text="3. Clone Params", command=self.clone_parameters, 
                             bg=self.colors['success'], fg='white', width=15, **action_btn_style)
        clone_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        export_btn = tk.Button(button_frame, text="📊 Export Excel", command=self.export_excel_report, 
                              bg='#9C27B0', fg='white', width=15, **action_btn_style)
        export_btn.pack(side=tk.LEFT, padx=(0, 8))
        
        # Parameter Configuration Section
        self.param_config_frame = tk.LabelFrame(main_frame, text="📋 Parameter Configuration", 
                                               font=('Segoe UI', 10, 'bold'),
                                               bg=self.colors['panel'], 
                                               fg=self.colors['text'],
                                               relief=tk.FLAT, 
                                               borderwidth=2)
        self.param_config_frame.grid(row=5, column=0, columnspan=3, sticky='nsew', pady=(15, 0), padx=2)
        main_frame.rowconfigure(5, weight=1)  # Give weight to param config
        
        self.parameters_header = tk.Label(
            self.param_config_frame, 
            text="⏳ Load files to see parameters...",
            font=('Segoe UI', 10),
            bg=self.colors['panel'],
            fg=self.colors['text_secondary']
        )
        self.parameters_header.pack(pady=15)
        
        # Results Section
        results_frame = tk.LabelFrame(main_frame, text="📊 Processing Results", 
                                     font=('Segoe UI', 10, 'bold'),
                                     bg=self.colors['panel'], 
                                     fg=self.colors['text'],
                                     relief=tk.FLAT, 
                                     borderwidth=2)
        results_frame.grid(row=6, column=0, columnspan=3, sticky='nsew', pady=(20, 10), padx=2)
        main_frame.rowconfigure(6, weight=3)  # Give more weight to results for scrolling
        
        text_frame = tk.Frame(results_frame, bg=self.colors['panel'])
        text_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        
        self.results_text = tk.Text(text_frame, height=8, wrap=tk.WORD,
                                   font=('Consolas', 9),
                                   relief=tk.SOLID,
                                   borderwidth=1,
                                   bg='#fafafa',
                                   fg=self.colors['text'])
        scrollbar = tk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        
        self.results_text.grid(row=0, column=0, sticky='nsew')
        scrollbar.grid(row=0, column=1, sticky='ns')
        
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
        self.update_results(
            "Welcome to Multi-File Parameter Cloner!\n\n"
            "Please follow the steps:\n"
            "1. Select parameter list file and add parameter files, then click 'Load Files'\n"
            "2. Define suffix array (up to 6 suffixes) and click 'Update Suffix Array'\n"
            "   - All non-empty suffixes will be used to create multiple parameters\n"
            "   - Example: 'Sport_1', 'Sport_2', 'Sport_3' → creates 3 new parameters\n"
            "3. Select cloning option and configure it:\n"
            "   - Add new suffix: Choose Prefix/Suffix/Replace position\n"
            "     * Each parameter will be cloned with ALL defined suffixes\n"
            "     * Replace option: Choose 'Replace directly' or 'Clone and replace'\n"
            "   - Replace suffix and split: Find existing suffix and create new parameter with new suffix\n"
            "4. Click 'Clone Parameters' to update all files"
        )
    
    def log_message(self, message):
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.results_text.insert(tk.END, f"{timestamp} - {message}\n")
        self.results_text.see(tk.END)
        self.root.update()
    
    def add_parameter_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Parameter Files",
            filetypes=[("All files", "*.*")]
        )
        
        if file_paths:
            for path in file_paths:
                if path not in self.parameter_files:
                    self.parameter_files.append(path)
                    self.file_listbox.insert(tk.END, os.path.basename(path))
            self.update_file_count()
    
    def scan_directory_for_parameters(self):
        directory = filedialog.askdirectory(title="Select Directory to Scan")
        if not directory:
            return
            
        self.list_file_path = self.list_file_var.get().strip()
        if not self.list_file_path or not os.path.isfile(self.list_file_path):
            messagebox.showerror("Error", "Please select a valid parameter list file first")
            return
        
        # Save directory to history
        if directory not in self.scan_directory_history:
            self.scan_directory_history.insert(0, directory)
            self.scan_directory_history = self.scan_directory_history[:10]  # Keep last 10
            self.save_config()
            
        self.start_scan_process([directory])

    def scan_multiple_directories(self):
        dialog = MultiDirectoryDialog(self.root)
        self.root.wait_window(dialog.dialog)
        
        if not dialog.selected_directories:
            return
            
        self.list_file_path = self.list_file_var.get().strip()
        if not self.list_file_path or not os.path.isfile(self.list_file_path):
            messagebox.showerror("Error", "Please select a valid parameter list file first")
            return
        
        self.start_scan_process(dialog.selected_directories)
        
    def start_scan_process(self, directories):
        try:
            self.parameter_list = read_parameter_list(self.list_file_path)
            if not self.parameter_list:
                messagebox.showerror("Error", "No parameters found in list file")
                return
        except Exception as e:
            messagebox.showerror("Error", f"Error reading parameter list: {str(e)}")
            return
        
        progress_window = tk.Toplevel(self.root)
        progress_window.title("Scanning Files")
        progress_window.geometry("500x200")
        progress_window.transient(self.root)
        
        progress_window.update_idletasks()
        x = (progress_window.winfo_screenwidth() // 2) - (progress_window.winfo_width() // 2)
        y = (progress_window.winfo_screenheight() // 2) - (progress_window.winfo_height() // 2)
        progress_window.geometry(f"+{x}+{y}")
        
        tk.Label(progress_window, text=f"Scanning {len(directories)} director{'y' if len(directories)==1 else 'ies'} for parameters...", 
                font=('Arial', 10, 'bold')).pack(pady=10)
        
        progress_var = tk.DoubleVar()
        progress_bar = ttk.Progressbar(progress_window, variable=progress_var, maximum=100)
        progress_bar.pack(fill="x", padx=20, pady=10)
        
        status_label = tk.Label(progress_window, text="Initializing...", font=('Arial', 9))
        status_label.pack(pady=5)
        
        dir_label = tk.Label(progress_window, text="", font=('Arial', 9), fg="blue")
        dir_label.pack(pady=5)
        
        def scan_thread():
            found_files = []
            total_files = 0
            processed_files = 0
            
            for directory in directories:
                dir_label.config(text=f"Counting files in: {os.path.basename(directory)}")
                progress_window.update_idletasks()
                for root, dirs, files in os.walk(directory):
                    total_files += len(files)
            
            for directory in directories:
                dir_label.config(text=f"Scanning in: {os.path.basename(directory)}")
                progress_window.update_idletasks()
                
                for root, dirs, files in os.walk(directory):
                    for file_name in files:
                        file_path = os.path.join(root, file_name)
                        processed_files += 1
                        
                        progress = (processed_files / total_files) * 100 if total_files > 0 else 0
                        progress_var.set(progress)
                        status_label.config(text=f"Scanning: {file_name}")
                        progress_window.update_idletasks()
                        
                        try:
                            if os.path.getsize(file_path) > 10_000_000:
                                continue
                        except:
                            continue
                            
                        try:
                            with open(file_path, 'r', errors='ignore') as f:
                                content = f.read()
                                
                            for param in self.parameter_list:
                                if any(pattern.search(content) for pattern in [
                                    re.compile(r'KENNLINIE\s+' + re.escape(param) + r'\s+'),
                                    re.compile(r'KENNFELD\s+' + re.escape(param) + r'\s+'),
                                    re.compile(r'FESTWERT\s+' + re.escape(param) + r'\s+'),
                                    re.compile(r'GRUPPENKENNLINIE\s+' + re.escape(param) + r'\s+')
                                ]):
                                    if file_path not in found_files:
                                        found_files.append(file_path)
                                    break
                                    
                        except Exception:
                            continue
           
            # Update UI with results
            def update_ui():
                progress_window.destroy()
                
                if not found_files:
                    messagebox.showinfo("Scan Complete", "No files containing the parameters were found.")
                    return
                    
                # Add the found files to our parameter files list
                for file_path in found_files:
                    if file_path not in self.parameter_files:
                        self.parameter_files.append(file_path)
                        self.file_listbox.insert(tk.END, os.path.basename(file_path))
                
                self.update_file_count()
                self.log_message(f"Scan complete: Found {len(found_files)} files containing parameters from the list.")
                
            # Schedule the UI update on the main thread
            self.root.after(0, update_ui)
        
        # Start the scanning thread
        threading.Thread(target=scan_thread, daemon=True).start()
    
    def remove_selected_file(self):
        selected_indices = self.file_listbox.curselection()
        
        # Remove from bottom to top to avoid index shifting issues
        for index in sorted(selected_indices, reverse=True):
            if index < len(self.parameter_files):
                file_path = self.parameter_files[index]
                del self.parameter_files[index]
                self.file_listbox.delete(index)
                
                # Remove from file_contents if it exists
                if file_path in self.file_contents:
                    del self.file_contents[file_path]
        
        self.update_file_count()

    def clear_all_files(self):
        self.parameter_files = []
        self.file_contents = {}
        self.file_listbox.delete(0, tk.END)
        self.update_file_count()
    
    def update_file_count(self):
        """Update the file count label."""
        count = len(self.parameter_files)
        self.file_count_label.config(text=f"Selected files: {count}")
    
    def select_list_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Parameter List File",
            filetypes=[("Text files", "*.txt"), ("All files", "*.*")]
        )
        
        if file_path:
            self.list_file_path = file_path
            self.list_file_var.set(file_path)
            
            # Save to history
            if file_path not in self.list_file_history:
                self.list_file_history.insert(0, file_path)
                self.list_file_history = self.list_file_history[:10]  # Keep last 10
                self.list_file_combo['values'] = self.list_file_history
                self.save_config()
    
    def on_position_change(self, *args):
        """Show/hide replace mode options based on position selection"""
        if hasattr(self, 'replace_mode_frame_suffix'):
            if self.position_var.get() == "Replace":
                self.replace_mode_frame_suffix.pack(fill='x', padx=5, pady=5)
            else:
                self.replace_mode_frame_suffix.pack_forget()
    
    def toggle_clone_options(self):
        option = self.clone_option_var.get()
        
        # Hide all option frames
        self.add_suffix_frame.pack_forget()
        self.replace_text_frame.pack_forget()
        
        # Show the selected option frame
        if option == "add_suffix":
            self.add_suffix_frame.pack(fill="x", padx=5, pady=5)
            self.on_position_change()  # Update replace mode visibility
        elif option == "replace_text":
            self.replace_text_frame.pack(fill="x", padx=5, pady=5)
        
        # Update parameter input labels based on the selected option
        if hasattr(self, 'parameter_widgets') and self.parameter_widgets:
            self.update_parameter_labels()
                
    def update_parameter_labels(self):
        option = self.clone_option_var.get()
        
        # Rebuild parameter config to update labels
        if hasattr(self, 'parameter_widgets') and self.parameter_widgets and self.file_contents:
            self.create_parameter_config_widgets()

    def load_files(self):
        # Check if parameter files are selected
        if not self.parameter_files:
            messagebox.showerror("Error", "Please select parameter files first")
            return
            
        # Check if parameter list file is selected
        self.list_file_path = self.list_file_var.get().strip()
        if not self.list_file_path or not os.path.isfile(self.list_file_path):
            messagebox.showerror("Error", "Please select a valid parameter list file")
            return
        
        # Save list file to history
        if self.list_file_path not in self.list_file_history:
            self.list_file_history.insert(0, self.list_file_path)
            self.list_file_history = self.list_file_history[:10]  # Keep last 10
            self.list_file_combo['values'] = self.list_file_history
            self.save_config()
        
        # Load parameter list
        try:
            self.parameter_list = read_parameter_list(self.list_file_path)
            if not self.parameter_list:
                messagebox.showerror("Error", "No parameters found in list file")
                return
        except Exception as e:
            messagebox.showerror("Error", f"Error reading parameter list: {str(e)}")
            return
        
        # Load file contents
        self.file_contents = {}
        failed_files = []
        
        for file_path in self.parameter_files:
            try:
                self.file_contents[file_path] = read_file(file_path)
            except Exception as e:
                failed_files.append(f"{os.path.basename(file_path)}: {str(e)}")
        
        if failed_files:
            messagebox.showwarning("Warning", f"Failed to read some files:\n" + "\n".join(failed_files))
        
        # Rebuild parameter configuration section
        self.create_parameter_config_widgets()
        
        self.log_message(f"Loaded {len(self.file_contents)} files and {len(self.parameter_list)} parameters.")
    
    def create_parameter_config_widgets(self):
        # Clear existing parameter configuration
        for widget in self.param_config_frame.winfo_children():
            widget.destroy()
        
        # Create scrollable frame for parameters
        canvas = tk.Canvas(self.param_config_frame, height=200)
        scrollbar = tk.Scrollbar(self.param_config_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas)
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Create header
        header_frame = tk.Frame(scrollable_frame)
        header_frame.pack(fill="x", padx=5, pady=5)
        
        tk.Label(header_frame, text="Parameter", width=25, font=('Arial', 9, 'bold'), 
                relief='ridge', bg='lightgray').grid(row=0, column=0, sticky="ew", padx=1)
        
        # The label text depends on the selected clone option
        if self.clone_option_var.get() == "add_suffix":
            label_text = "Suffix Count"
        else:
            label_text = "Will process if contains suffix:"
        tk.Label(header_frame, text=label_text, width=20, font=('Arial', 9, 'bold'), 
                relief='ridge', bg='lightgray').grid(row=0, column=1, sticky="ew", padx=1)
        tk.Label(header_frame, text="Found In", width=12, font=('Arial', 9, 'bold'), 
                relief='ridge', bg='lightgray').grid(row=0, column=2, sticky="ew", padx=1)
        
        self.parameter_widgets = {}
        
        # Get available suffixes
        self.update_suffix_array()
        available_suffixes = []
        for entry in self.suffix_entries:
            value = entry.get().strip()
            if value:
                available_suffixes.append(value)
        
        # Create row for each parameter
        for i, param in enumerate(self.parameter_list):
            param_frame = tk.Frame(scrollable_frame)
            param_frame.pack(fill="x", pady=1, padx=5)
            
            # Parameter name with alternating row colors
            bg_color = 'white' if i % 2 == 0 else '#f0f0f0'
            tk.Label(param_frame, text=param, width=25, font=('Arial', 9), 
                    relief='ridge', bg=bg_color, anchor='w').grid(row=0, column=0, sticky="ew", padx=1)
            
            # Show suffix count or processing indicator
            if self.clone_option_var.get() == "add_suffix":
                # Count non-empty suffixes
                suffix_count = len([e.get().strip() for e in self.suffix_entries if e.get().strip()])
                suffix_widget = tk.Label(param_frame, text=f"{suffix_count} suffix(es) will be created", 
                                        font=('Arial', 9), relief='ridge', bg='lightyellow')
            else:
                # For replace_text mode, show checkbox or indicator
                suffix_widget = tk.Label(param_frame, text="Will process", 
                                        font=('Arial', 9), relief='ridge', bg='lightblue')
            
            suffix_widget.grid(row=0, column=1, sticky="ew", padx=1)
            
            # Count how many files this parameter appears in
            found_count = 0
            for content in self.file_contents.values():
                if find_specific_kennlinie(content, param):
                    found_count += 1
            
            count_color = 'lightgreen' if found_count > 0 else 'lightcoral'
            tk.Label(param_frame, text=f"{found_count} file(s)", width=12, font=('Arial', 9), 
                    relief='ridge', bg=count_color).grid(row=0, column=2, sticky="ew", padx=1)
            
            # Store the widget reference
            self.parameter_widgets[param] = suffix_widget
        
        canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        scrollbar.pack(side="right", fill="y")
        
        # Bind mousewheel to canvas
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def update_suffix_array(self):
        # Update suffix array from the entry widgets
        self.suffix_array = []
        for entry in self.suffix_entries:
            value = entry.get().strip()
            if value:
                self.suffix_array.append(value)
        
        # Rebuild parameter widgets to show updated suffix count
        if hasattr(self, 'parameter_widgets') and self.parameter_widgets and self.file_contents:
            self.create_parameter_config_widgets()
        
        self.log_message(f"Updated suffix array: {', '.join(self.suffix_array)} ({len(self.suffix_array)} suffix(es))")

    def clone_parameters(self):
        # Check if files and parameters are loaded
        if not self.parameter_files or not self.parameter_list or not self.file_contents:
            messagebox.showerror("Error", "Please load files and parameters first")
            return
        
        # Confirm before proceeding
        if not messagebox.askyesno("Confirm Clone", 
            f"Clone parameters in {len(self.parameter_files)} files?\n\n"
            "This action will modify the original files."):
            return
        
        # Clear cloning history for new operation
        self.cloning_history = []
        
        # Get the clone option
        clone_option = self.clone_option_var.get()
        
        # Process each file
        modified_count = 0
        total_clones = 0
        
        self.log_message("Starting parameter cloning process...")
        
        for file_path, content in self.file_contents.items():
            file_modified = False
            updated_content = content
            file_clones = 0
            
            self.log_message(f"\nProcessing file: {os.path.basename(file_path)}")
            
            # Process each parameter in the file
            for param in self.parameter_list:
                if clone_option == "add_suffix":
                    # Option 1: Add suffix
                    position = self.position_var.get()
                    
                    # Get all non-empty suffixes from suffix array
                    suffixes_to_use = []
                    for entry in self.suffix_entries:
                        suffix_value = entry.get().strip()
                        if suffix_value:
                            suffixes_to_use.append(suffix_value)
                    
                    if not suffixes_to_use:
                        continue  # Skip if no suffix provided
                    
                    # Process with each suffix from suffix array
                    for suffix in suffixes_to_use:
                            if position == "Replace":
                                # Replace mode: need to check replace_mode_suffix_var
                                replace_mode = self.replace_mode_suffix_var.get()
                                
                                # Get the original block
                                block = find_specific_kennlinie(updated_content, param)
                                if not block:
                                    continue
                                
                                # Create new block with replaced name
                                lines = block.split('\n')
                                first_line_parts = lines[0].split(maxsplit=2)
                                
                                if len(first_line_parts) < 2:
                                    continue
                                
                                param_type = first_line_parts[0]
                                
                                # Handle underscore intelligently
                                if param.endswith('_') or suffix.startswith('_'):
                                    new_name = f"{param}{suffix}"  # No extra underscore needed
                                else:
                                    new_name = suffix  # Just use suffix as is
                                
                                new_first_line = f"{param_type} {new_name}"
                                
                                if len(first_line_parts) > 2:
                                    new_first_line += f" {first_line_parts[2]}"
                                    
                                lines[0] = new_first_line
                                
                                # Clean the lines
                                cleaned_lines = [lines[0]]
                                for line in lines[1:]:
                                    if line.strip():
                                        cleaned_lines.append(line)
                                
                                new_block = '\n'.join(cleaned_lines).strip()
                                
                                # Check if new parameter already exists
                                existing_new_param = find_specific_kennlinie(updated_content, new_name)
                                
                                if replace_mode == "direct_replace":
                                    # Replace directly: modify the existing parameter in place
                                    updated_content = updated_content.replace(block, new_block)
                                    file_modified = True
                                    file_clones += 1
                                    self.log_message(f"  - Replaced {param} with {suffix} (direct)")
                                    
                                    # Track for Excel report - store full blocks
                                    self.cloning_history.append({
                                        'file': file_path,
                                        'folder': os.path.dirname(file_path),
                                        'old_param': param,
                                        'new_param': new_name,
                                        'old_value': block,  # Full old block
                                        'new_value': new_block,  # Full new block
                                        'type': param_type,
                                        'operation': 'Replace (Direct)',
                                        'status': 'Success'
                                    })
                                elif existing_new_param:
                                    # New parameter already exists - skip cloning
                                    self.log_message(f"  - Skipped {new_name}: Already Splitted")
                                    
                                    # Track as already split
                                    self.cloning_history.append({
                                        'file': file_path,
                                        'folder': os.path.dirname(file_path),
                                        'old_param': param,
                                        'new_param': new_name,
                                        'old_value': block,
                                        'new_value': existing_new_param,
                                        'type': param_type,
                                        'operation': 'Replace (Clone)',
                                        'status': 'Already Splitted'
                                    })
                                else:
                                    # Clone and replace: create new parameter with new name
                                    updated_content = updated_content.rstrip() + "\n\n" + new_block + "\n"
                                    file_modified = True
                                    file_clones += 1
                                    self.log_message(f"  - Cloned {param} as {suffix} (new parameter)")
                                    
                                    # Track for Excel report - store full blocks
                                    self.cloning_history.append({
                                        'file': file_path,
                                        'folder': os.path.dirname(file_path),
                                        'old_param': param,
                                        'new_param': new_name,
                                        'old_value': block,  # Full old block
                                        'new_value': new_block,  # Full new block
                                        'type': param_type,
                                        'operation': 'Replace (Clone)',
                                        'status': 'Success'
                                    })
                            else:
                                # Prefix or Suffix mode: just clone normally
                                cloned_block = clone_parameter(updated_content, param, suffix, position)
                                
                                if cloned_block:
                                    # Determine new parameter name
                                    if position == "Prefix":
                                        if suffix.endswith('_') or param.startswith('_'):
                                            new_param_name = f"{suffix}{param}"
                                        else:
                                            new_param_name = f"{suffix}_{param}"
                                    else:  # Suffix
                                        if param.endswith('_') or suffix.startswith('_'):
                                            new_param_name = f"{param}{suffix}"
                                        else:
                                            new_param_name = f"{param}_{suffix}"
                                    
                                    # Check if new parameter already exists
                                    existing_new_param = find_specific_kennlinie(updated_content, new_param_name)
                                    
                                    if existing_new_param:
                                        # Parameter already exists - skip cloning
                                        position_text = "prefix" if position == "Prefix" else "suffix"
                                        self.log_message(f"  - Skipped {new_param_name}: Already Splitted")
                                        
                                        # Track as already split
                                        old_block = find_specific_kennlinie(content, param)
                                        self.cloning_history.append({
                                            'file': file_path,
                                            'folder': os.path.dirname(file_path),
                                            'old_param': param,
                                            'new_param': new_param_name,
                                            'old_value': old_block if old_block else "",
                                            'new_value': existing_new_param,
                                            'type': extract_parameter_type(cloned_block),
                                            'operation': f'{position_text.capitalize()}',
                                            'status': 'Already Splitted'
                                        })
                                    else:
                                        # Add the cloned block to the end of the file
                                        updated_content = updated_content.rstrip() + "\n\n" + cloned_block + "\n"
                                        file_modified = True
                                        file_clones += 1
                                        position_text = "prefix" if position == "Prefix" else "suffix"
                                        self.log_message(f"  - Cloned {param} with {position_text} '{suffix}'")
                                        
                                        # Track for Excel report - store full blocks
                                        old_block = find_specific_kennlinie(content, param)
                                        
                                        self.cloning_history.append({
                                            'file': file_path,
                                            'folder': os.path.dirname(file_path),
                                            'old_param': param,
                                            'new_param': new_param_name,
                                            'old_value': old_block if old_block else "",  # Full old block
                                            'new_value': cloned_block,  # Full new block
                                            'type': extract_parameter_type(cloned_block),
                                            'operation': f'{position_text.capitalize()}',
                                            'status': 'Success'
                                        })
                
                elif clone_option == "replace_text":
                    # Option 2: Replace text in parameter name (always creates new parameter)
                    find_text = self.find_text_entry.get().strip()
                    replace_with = self.replace_with_entry.get().strip()
                    
                    if not find_text or not replace_with:
                        continue
                    
                    # Check if this parameter contains the text to replace
                    if find_text in param:
                        # Create the new parameter name
                        new_param_name = param.replace(find_text, replace_with)
                        
                        # Get the parameter block from the file content
                        block = find_specific_kennlinie(updated_content, param)
                        
                        if not block:
                            continue
                        
                        # Create the new parameter block by replacing the parameter name
                        lines = block.split('\n')
                        first_line_parts = lines[0].split(maxsplit=2)
                        
                        if len(first_line_parts) < 2:
                            continue
                        
                        param_type = first_line_parts[0]
                        new_first_line = f"{param_type} {new_param_name}"
                        
                        if len(first_line_parts) > 2:
                            new_first_line += f" {first_line_parts[2]}"
                            
                        lines[0] = new_first_line
                        
                        # Clean the lines
                        cleaned_lines = [lines[0]]
                        for line in lines[1:]:
                            if line.strip():
                                cleaned_lines.append(line)
                        
                        new_block = '\n'.join(cleaned_lines).strip()
                        
                        # Check if new parameter already exists
                        existing_new_param = find_specific_kennlinie(updated_content, new_param_name)
                        
                        if existing_new_param:
                            # Parameter already exists - skip cloning
                            self.log_message(f"  - Skipped {new_param_name}: Already Splitted")
                            
                            # Track as already split
                            self.cloning_history.append({
                                'file': file_path,
                                'folder': os.path.dirname(file_path),
                                'old_param': param,
                                'new_param': new_param_name,
                                'old_value': block,
                                'new_value': existing_new_param,
                                'type': param_type,
                                'operation': 'Replace Text',
                                'status': 'Already Splitted'
                            })
                        else:
                            # Always create new parameter (clone mode)
                            updated_content = updated_content.rstrip() + "\n\n" + new_block + "\n"
                            file_modified = True
                            file_clones += 1
                            self.log_message(f"  - Created {new_param_name} from {param}")
                            
                            # Track for Excel report - store full blocks
                            self.cloning_history.append({
                                'file': file_path,
                                'folder': os.path.dirname(file_path),
                                'old_param': param,
                                'new_param': new_param_name,
                                'old_value': block,  # Full old block
                                'new_value': new_block,  # Full new block
                                'type': param_type,
                                'operation': 'Replace Text',
                                'status': 'Success'
                            })
            
            # Save the updated file if modified
            if file_modified:
                try:
                    write_file(file_path, updated_content)
                    modified_count += 1
                    total_clones += file_clones
                    self.log_message(f"  Saved {file_clones} cloned parameters")
                except Exception as e:
                    messagebox.showerror("Error", f"Error writing to file {os.path.basename(file_path)}: {str(e)}")
            else:
                self.log_message("  No parameters found to clone")
        
        # Final summary
        if modified_count > 0:
            summary_msg = f"Clone operation completed successfully!\n\n" \
                         f"Modified files: {modified_count}\n" \
                         f"Total parameters cloned: {total_clones}"
            self.log_message(f"\n{summary_msg}")
            messagebox.showinfo("Success", summary_msg)
        else:
            self.log_message("\nNo files were modified. No parameters matched the criteria.")
            messagebox.showinfo("Info", "No files were modified. No parameters matched the criteria.")

    def generate_sheet_current_parameters(self, wb, file_params_cache):
        """Generate Sheet 1: Current Parameters with Values - Only checking list parameters"""
        ws = wb.create_sheet("Current Parameters", 0)
        
        # Define header style
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers
        headers = ["Parameter Name", "Value", "File Name", "Folder Path", "Status"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 35  # Parameter Name
        ws.column_dimensions['B'].width = 60  # Value (entire block)
        ws.column_dimensions['C'].width = 30  # File Name
        ws.column_dimensions['D'].width = 45  # Folder Path
        ws.column_dimensions['E'].width = 12  # Status
        
        # Populate data - only for parameters in the checking list
        row_num = 2
        found_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
        missing_font = Font(color="9C0006", bold=True)  # Dark red text
        
        # Iterate through each parameter in the checking list
        for param_name in self.parameter_list:
            # Check each file for this parameter
            for file_path, all_params in file_params_cache.items():
                file_name = os.path.basename(file_path)
                folder_path = os.path.dirname(file_path)
                
                # Look for the parameter in this file's parameters
                param_found = False
                param_block = None
                
                for param_detail in all_params:
                    if param_detail['name'] == param_name:
                        param_found = True
                        param_block = param_detail['block']
                        break
                
                # Prepare row data
                if param_found:
                    # Parameter exists in this file
                    row_data = [param_name, param_block, file_name, folder_path, "Found"]
                    ws.append(row_data)
                    
                    # Apply green formatting for found parameters
                    for cell in ws[row_num]:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.fill = found_fill
                else:
                    # Parameter missing in this file
                    row_data = [param_name, "NA", file_name, folder_path, "Missing"]
                    ws.append(row_data)
                    
                    # Apply red formatting for missing parameters
                    for cell in ws[row_num]:
                        cell.border = border
                        cell.alignment = Alignment(vertical="top", wrap_text=True)
                        cell.fill = missing_fill
                        cell.font = missing_font
                
                row_num += 1
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def generate_sheet_folder_structure(self, wb, file_params_cache):
        """Generate Sheet 2: Folder and File Structure"""
        ws = wb.create_sheet("Folder Structure", 1)
        
        # Define styles
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        folder_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers
        headers = ["Folder Path", "File Name", "Full File Path", "Parameters Found", "Parameter Count"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 45
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 40
        ws.column_dimensions['E'].width = 15
        
        # Group files by folder
        folder_dict = {}
        for file_path in file_params_cache.keys():
            folder_path = os.path.dirname(file_path)
            if folder_path not in folder_dict:
                folder_dict[folder_path] = []
            folder_dict[folder_path].append(file_path)
        
        # Populate data using cached parameters
        row_num = 2
        for folder_path in sorted(folder_dict.keys()):
            for file_path in sorted(folder_dict[folder_path]):
                file_name = os.path.basename(file_path)
                all_params = file_params_cache[file_path]
                param_names = [p['name'] for p in all_params]
                param_count = len(param_names)
                params_str = ", ".join(param_names[:5])  # Show first 5
                if param_count > 5:
                    params_str += f"... (+{param_count - 5} more)"
                
                row_data = [folder_path, file_name, file_path, params_str, param_count]
                ws.append(row_data)
                
                # Apply formatting
                for cell in ws[row_num]:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if cell.column == 1:  # Folder column
                        cell.fill = folder_fill
                        cell.font = Font(bold=True)
                
                row_num += 1
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def generate_sheet_cloning_comparison(self, wb):
        """Generate Sheet 3: Split Comparison (Old vs New)"""
        ws = wb.create_sheet("Split Comparison", 2)
        
        # Define styles
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="000000", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        new_param_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Blue
        replace_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow
        already_split_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")  # Gray
        already_split_font = Font(color="7F7F7F", italic=True)  # Gray italic text
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Set headers - removed Old Type and New Type columns
        headers = ["Folder", "File", "Old Parameter", "Old Value", 
                   "New Parameter", "New Value", "Operation", "Status"]
        ws.append(headers)
        
        # Style header row
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 40  # Folder
        ws.column_dimensions['B'].width = 25  # File
        ws.column_dimensions['C'].width = 30  # Old Parameter
        ws.column_dimensions['D'].width = 60  # Old Value (full block)
        ws.column_dimensions['E'].width = 30  # New Parameter
        ws.column_dimensions['F'].width = 60  # New Value (full block)
        ws.column_dimensions['G'].width = 18  # Operation
        ws.column_dimensions['H'].width = 12  # Status
        
        # Populate data from cloning history
        row_num = 2
        for entry in self.cloning_history:
            folder = entry.get('folder', '')
            file_name = os.path.basename(entry.get('file', ''))
            old_param = entry.get('old_param', '')
            old_value = entry.get('old_value', '')  # Full block
            new_param = entry.get('new_param', '')
            new_value = entry.get('new_value', '')  # Full block
            operation = entry.get('operation', '')
            status = entry.get('status', '')
            
            row_data = [folder, file_name, old_param, old_value,
                       new_param, new_value, operation, status]
            ws.append(row_data)
            
            # Apply formatting
            for cell in ws[row_num]:
                cell.border = border
                cell.alignment = Alignment(vertical="top", wrap_text=True)  # Top alignment for blocks
                
                # Color code based on status and operation
                if status == "Already Splitted":
                    cell.fill = already_split_fill
                    cell.font = already_split_font
                elif "Replace" in operation:
                    cell.fill = replace_fill
                else:
                    cell.fill = new_param_fill
            
            row_num += 1
        
        # If no cloning history, show message
        if not self.cloning_history:
            ws.append(["No cloning operations recorded yet.", "", "", "", "", "", "", ""])
            ws['A2'].font = Font(italic=True, color="666666")
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def export_excel_report(self):
        """Export comprehensive Excel report with 3 sheets"""
        # Check if openpyxl is available
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Error", 
                "openpyxl library is not installed.\n\n"
                "Please install it using:\npip install openpyxl")
            return
        
        # Check if data is loaded
        if not self.file_contents or not self.parameter_list:
            messagebox.showerror("Error", 
                "Please load files and parameters first before exporting report.")
            return
        
        try:
            self.log_message("Starting Excel report generation...")
            self.root.update()  # Update UI
            
            # Pre-extract all parameters from all files (cache for reuse)
            self.log_message(f"Analyzing {len(self.file_contents)} files...")
            self.root.update()
            
            file_params_cache = {}
            for idx, (file_path, content) in enumerate(self.file_contents.items(), 1):
                file_params_cache[file_path] = extract_all_parameters(content)
                if idx % 10 == 0:  # Log progress every 10 files
                    self.log_message(f"  Processed {idx}/{len(self.file_contents)} files...")
                    self.root.update()
            
            total_params = sum(len(params) for params in file_params_cache.values())
            self.log_message(f"Found {total_params} parameters across all files")
            self.root.update()
            
            # Create Report folder if it doesn't exist
            report_folder = os.path.join(os.path.dirname(__file__), 'Report')
            if not os.path.exists(report_folder):
                os.makedirs(report_folder)
                self.log_message(f"Created Report folder: {report_folder}")
            
            # Create workbook
            wb = Workbook()
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Generate all sheets with cached data
            self.log_message("Generating Sheet 1: Current Parameters...")
            self.root.update()
            self.generate_sheet_current_parameters(wb, file_params_cache)
            
            self.log_message("Generating Sheet 2: Folder Structure...")
            self.root.update()
            self.generate_sheet_folder_structure(wb, file_params_cache)
            
            self.log_message("Generating Sheet 3: Cloning Comparison...")
            self.root.update()
            self.generate_sheet_cloning_comparison(wb)
            
            # Generate filename with timestamp
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            excel_filename = f"Parameter_Report_{timestamp}.xlsx"
            excel_path = os.path.join(report_folder, excel_filename)
            
            # Save workbook
            self.log_message("Saving Excel file...")
            self.root.update()
            wb.save(excel_path)
            
            self.log_message(f"Excel report saved successfully!")
            self.log_message(f"Location: {excel_path}")
            
            # Show success message
            result = messagebox.showinfo("Success", 
                f"Excel report generated successfully!\n\n"
                f"File: {excel_filename}\n"
                f"Location: {report_folder}\n\n"
                f"Total files analyzed: {len(self.file_contents)}\n"
                f"Parameters tracked: {len(self.parameter_list)}\n"
                f"Cloning operations: {len(self.cloning_history)}")
            
            # Optional: Open folder in Windows Explorer
            if messagebox.askyesno("Open Folder", "Do you want to open the Report folder?"):
                os.startfile(report_folder)
                
        except Exception as e:
            error_msg = f"Error generating Excel report: {str(e)}"
            self.log_message(error_msg)
            messagebox.showerror("Export Error", error_msg)

    def load_config(self):
        """Load configuration including path histories from config file"""
        try:
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
                    self.list_file_history = config.get('list_file_history', [])
                    self.scan_directory_history = config.get('scan_directory_history', [])
        except Exception as e:
            print(f"Error loading config: {e}")
            self.list_file_history = []
            self.scan_directory_history = []
    
    def save_config(self):
        """Save configuration including path histories to config file"""
        try:
            config = {}
            if os.path.exists(self.config_file):
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    config = json.load(f)
            
            config['list_file_history'] = self.list_file_history
            config['scan_directory_history'] = self.scan_directory_history
            
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error saving config: {e}")

    def update_results(self, message):
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(tk.END, message)
        self.results_text.see(tk.END)

# Run the application
if __name__ == "__main__":
    root = tk.Tk()
    app = ParameterClonerApp(root)
    root.mainloop()